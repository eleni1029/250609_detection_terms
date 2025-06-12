#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
script_01_generate_xlsx.py (v2.5 - 包容關係優先處理版本)

新增功能：
1. 檢測敏感詞之間的包容關係
2. 根據包容關係確定優先順序
3. 按優先順序進行匹配，避免重複檢測被包容詞

修復內容：
1. 修復Excel語言區塊解析邏輯，正確處理合併儲存格
2. 改善語言名稱檢測，避免將表頭誤認為語言
3. 增強錯誤處理和調試資訊
"""

import json
import re
import itertools
import sys
import shutil
import datetime
from pathlib import Path
from collections import defaultdict
from config_loader import get_config

try:
    import polib
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌ 缺少必要套件：{e}")
    print("請執行：pip install polib openpyxl")
    sys.exit(1)


class InclusionDetector:
    """處理敏感詞包容關係和優先順序的類"""
    
    def __init__(self, sensitive_words_dict):
        """
        初始化包容關係檢測器
        
        Args:
            sensitive_words_dict: 敏感詞字典 {category: {keyword: {business_type: replacement}}}
        """
        self.sensitive_words_dict = sensitive_words_dict
        self.flat_words = self._flatten_words()
        self.inclusion_relationships = self._detect_inclusions()
        self.priority_sorted_words = self._sort_by_priority()
        
        # 調試輸出
        self._print_analysis()
    
    def _flatten_words(self):
        """將分層的敏感詞字典展平為 {keyword: word_info} 格式"""
        flat_words = {}
        for category, keywords in self.sensitive_words_dict.items():
            for keyword, business_replacements in keywords.items():
                flat_words[keyword] = {
                    'category': category,
                    'replacements': business_replacements,
                    'keyword': keyword
                }
        return flat_words
    
    def _detect_inclusions(self):
        """
        檢測敏感詞之間的包容關係
        
        Returns:
            dict: {包容詞: [被包容詞列表]}
        """
        inclusions = defaultdict(list)
        words = list(self.flat_words.keys())
        
        for i, word1 in enumerate(words):
            for j, word2 in enumerate(words):
                if i != j and word2 in word1 and len(word2) < len(word1):
                    inclusions[word1].append(word2)
        
        # 按被包容詞的長度排序（長的優先）
        for key in inclusions:
            inclusions[key].sort(key=len, reverse=True)
        
        return dict(inclusions)
    
    def _sort_by_priority(self):
        """
        根據包容關係確定優先順序
        
        Returns:
            list: 按優先順序排序的敏感詞列表
        """
        words = list(self.flat_words.keys())
        
        # 計算每個詞的優先級權重
        word_weights = {}
        
        for word in words:
            # 基礎權重 = 詞長度
            weight = len(word)
            
            # 如果該詞包容其他詞，增加權重
            if word in self.inclusion_relationships:
                weight += len(self.inclusion_relationships[word]) * 10
                
            # 如果該詞被其他詞包容，降低權重
            for parent_word, included_words in self.inclusion_relationships.items():
                if word in included_words:
                    weight -= 5
            
            word_weights[word] = weight
        
        # 按權重降序排序（權重高的優先）
        sorted_words = sorted(words, key=lambda w: word_weights[w], reverse=True)
        
        return sorted_words
    
    def _print_analysis(self):
        """輸出包容關係分析結果 - 簡化版"""
        inclusion_count = len(self.inclusion_relationships)
        total_words = len(self.flat_words)
        
        if inclusion_count > 0:
            print(f"   🔍 包容關係：{inclusion_count} 組，總詞數：{total_words}")
        else:
            print(f"   📝 總詞數：{total_words}（無包容關係）")
    
    def detect_with_priority(self, text, log_detail=None):
        """
        按優先順序檢測敏感詞，避免重複匹配被包容詞
        
        Args:
            text: 要檢測的文本
            log_detail: 日誌記錄函數（可選）
            
        Returns:
            list: 檢測到的敏感詞列表，每個元素包含 {keyword, category, replacements, positions}
        """
        detected_items = []
        processed_positions = set()  # 記錄已處理的字符位置
        
        for keyword in self.priority_sorted_words:
            word_info = self.flat_words[keyword]
            
            # 使用正則表達式查找所有匹配位置
            pattern = re.escape(keyword)
            matches = list(re.finditer(pattern, text))
            
            for match in matches:
                start_pos = match.start()
                end_pos = match.end()
                
                # 檢查該位置是否已被處理
                positions = set(range(start_pos, end_pos))
                if not positions.intersection(processed_positions):
                    # 記錄檢測結果
                    detected_items.append({
                        'keyword': keyword,
                        'category': word_info['category'],
                        'replacements': word_info['replacements'],
                        'start_pos': start_pos,
                        'end_pos': end_pos,
                        'matched_text': text[start_pos:end_pos]
                    })
                    
                    # 標記這些位置已處理
                    processed_positions.update(positions)
                    
                    # 只記錄到日誌，不打印到控制台
                    if log_detail:
                        log_detail(f"檢測到：「{keyword}」位置 {start_pos}-{end_pos}")
        
        return detected_items


def parse_language_blocks_from_excel(excel_path: Path, config):
    """
    修復版：解析語言獨立橫向分區塊 Excel，正確處理合併儲存格
    
    Args:
        excel_path: Excel 檔案路徑
        config: 配置物件
        
    Returns:
        dict: 每個語言的敏感詞和替換方案字典
    """
    
    print(f"📖 載入語言獨立橫向分區塊對照表：{excel_path.name}")
    
    # 載入工作簿
    wb = load_workbook(excel_path, data_only=True)
    
    # 獲取主工作表
    excel_config = config.get_excel_config()
    worksheet_name = excel_config.get('worksheets', {}).get('comparison', 'phrase_comparison')
    
    if worksheet_name not in wb.sheetnames:
        available_sheets = ', '.join(wb.sheetnames)
        raise ValueError(f"找不到工作表 '{worksheet_name}'，可用工作表：{available_sheets}")
    
    ws = wb[worksheet_name]
    
    # 獲取業態配置
    business_types = config.get_business_types()
    business_count = len(business_types)
    business_names = [bt_config['display_name'] for bt_config in business_types.values()]
    
    # 橫向配置
    horizontal_config = excel_config.get('horizontal_layout', {})
    block_separator = horizontal_config.get('block_separator_columns', 1)
    
    # 計算每個語言區塊的寬度：敏感詞類型 + 敏感詞 + 業態數量
    block_width = 2 + business_count
    
    language_data = {}
    warnings = []
    
    # 修復版：改進語言區塊檢測邏輯
    current_col = 1
    max_col = ws.max_column
    
    print(f"   Excel 最大列數：{max_col}")
    print(f"   每個區塊寬度：{block_width}")
    print(f"   區塊分隔：{block_separator}")
    
    while current_col <= max_col:
        # 檢查第1行是否有合併儲存格（語言標題）
        lang_cell = ws.cell(row=1, column=current_col)
        
        # 跳過空白儲存格
        if not lang_cell.value:
            current_col += 1
            continue
        
        language_name = str(lang_cell.value).strip()
        
        # 修復：排除表頭關鍵字，只接受真正的語言代碼
        excluded_headers = ['敏感詞類型', '敏感詞', '類型', 'type', 'keyword', 'category']
        if language_name.lower() in [h.lower() for h in excluded_headers]:
            print(f"   跳過表頭：{language_name} (列 {current_col})")
            current_col += 1
            continue
        
        # 修復：檢查是否是有效的語言代碼格式
        # 語言代碼通常是 xx_XX, xx-XX 或 xx 格式
        if not re.match(r'^[a-z]{2}([_-][A-Z]{2})?$', language_name):
            print(f"   跳過無效語言格式：{language_name} (列 {current_col})")
            current_col += 1
            continue
        
        print(f"   解析語言區塊：{language_name} (列 {current_col}-{current_col + block_width - 1})")
        
        # 檢查第2行的標題是否正確
        expected_headers = ["敏感詞類型", "敏感詞"] + business_names
        header_valid = True
        
        for i, expected_header in enumerate(expected_headers):
            col = current_col + i
            if col <= max_col:
                header_cell = ws.cell(row=2, column=col)
                actual_header = str(header_cell.value).strip() if header_cell.value else ""
                
                if actual_header != expected_header:
                    warnings.append(f"語言 {language_name} 區塊列 {col} 標題不符：期望 '{expected_header}'，實際 '{actual_header}'")
                    
                    # 如果基礎標題都不對，可能不是語言區塊
                    if i < 2 and actual_header not in ["敏感詞類型", "敏感詞"]:
                        header_valid = False
                        break
        
        if not header_valid:
            print(f"   跳過無效區塊：{language_name} (標題格式不符)")
            current_col += 1
            continue
        
        # 解析該語言的敏感詞和替換方案
        language_keywords = defaultdict(lambda: defaultdict(list))
        category_counts = defaultdict(int)
        
        # 從第3行開始讀取數據
        current_row = 3
        current_category = None
        
        while current_row <= ws.max_row:
            # 讀取敏感詞類型
            category_cell = ws.cell(row=current_row, column=current_col)
            category_value = str(category_cell.value).strip() if category_cell.value else ""
            
            if category_value:
                current_category = category_value
            
            # 讀取敏感詞
            keyword_cell = ws.cell(row=current_row, column=current_col + 1)
            keyword_value = str(keyword_cell.value).strip() if keyword_cell.value else ""
            
            # 如果沒有敏感詞，結束該語言區塊
            if not keyword_value:
                current_row += 1
                continue
            
            if not current_category:
                current_row += 1
                continue
            
            # 讀取各業態的替換方案
            business_replacements = {}
            
            for bt_index, (bt_code, bt_config) in enumerate(business_types.items()):
                col = current_col + 2 + bt_index
                if col <= max_col:
                    replacement_cell = ws.cell(row=current_row, column=col)
                    replacement_value = str(replacement_cell.value).strip() if replacement_cell.value else ""
                    
                    if replacement_value:
                        business_replacements[bt_code] = replacement_value
            
            # 儲存到語言數據中
            language_keywords[current_category][keyword_value] = business_replacements
            category_counts[current_category] += 1
            
            current_row += 1
            
            # 如果讀取了足夠多的行且沒有更多數據，退出
            if current_row > ws.max_row or current_row - 3 > 50:  # 限制最多讀50行
                break
        
        # 只有當找到有效數據時才加入結果
        if language_keywords:
            language_data[language_name] = dict(language_keywords)
            
            total_keywords = sum(category_counts.values())
            replacement_counts = {}
            
            for bt_code in business_types.keys():
                count = 0
                for category_data in language_keywords.values():
                    for keyword_data in category_data.values():
                        if bt_code in keyword_data:
                            count += 1
                replacement_counts[bt_code] = count
            
            print(f"     發現語言區塊：{language_name}")
            print(f"       {language_name}: {total_keywords} 個敏感詞")
            
            for category, count in category_counts.items():
                print(f"         {category}: {count} 個敏感詞")
                
            for bt_code, bt_config in business_types.items():
                count = replacement_counts.get(bt_code, 0)
                print(f"         {bt_config['display_name']}: {count} 個有替換方案")
        else:
            print(f"   語言區塊 {language_name} 未找到有效數據")
        
        # 移動到下個可能的語言區塊
        current_col += block_width + block_separator
    
    # 輸出警告
    if warnings:
        print("⚠️  解析警告：")
        for i, warning in enumerate(warnings[:30]):  # 限制顯示前30個警告
            print(f"     {warning}")
        if len(warnings) > 30:
            print(f"     ... 還有 {len(warnings) - 30} 個警告")
    
    # 修復：總結實際發現的語言
    if language_data:
        total_languages = len(language_data)
        total_categories = len(set().union(*[keywords.keys() for keywords in language_data.values()]))
        total_keywords = sum(sum(len(category.keys()) for category in keywords.values()) for keywords in language_data.values())
        
        print(f"✅ 成功載入 {total_languages} 個語言區塊")
        for language_name, keywords in language_data.items():
            keyword_count = sum(len(category.keys()) for category in keywords.values())
            category_count = len(keywords.keys())
            print(f"   {language_name}: {keyword_count} 個敏感詞，{category_count} 個分類")
            
            # 統計各業態的替換方案數量
            for bt_code, bt_config in business_types.items():
                count = 0
                for category_data in keywords.values():
                    for keyword_data in category_data.values():
                        if bt_code in keyword_data:
                            count += 1
                print(f"     {bt_config['display_name']}: {count} 個有替換方案")
    else:
        print("❌ 未找到任何有效的語言區塊")
    
    return language_data


def detect_sensitive_phrases_in_files_with_priority(config, language: str, sensitive_words: dict):
    """
    使用優先順序邏輯在指定語言的翻譯檔案中檢測敏感詞
    
    Args:
        config: 配置物件
        language: 語言代碼
        sensitive_words: 敏感詞字典 {category: {keyword: {business_type: replacement, ...}, ...}}
        
    Returns:
        list: 檢測到的敏感詞項目列表
    """
    
    print(f"   🔍 檢測敏感詞...")
    
    # 初始化包容關係檢測器
    detector = InclusionDetector(sensitive_words)
    
    detected_items = []
    
    # 創建日誌記錄函數
    def log_detail(message):
        # 這裡可以寫入日誌檔案，但不打印到控制台
        pass
    
    try:
        # 獲取語言檔案
        language_files = config.get_language_files(language)
        
        # 檢測 PO 檔案
        if 'po_file' in language_files:
            po_path = language_files['po_file']
            if po_path.exists():
                try:
                    po_data = polib.pofile(str(po_path))
                    
                    for entry in po_data:
                        if not entry.msgstr:  # 跳過未翻譯的項目
                            continue
                        
                        # 使用優先順序檢測
                        detected = detector.detect_with_priority(entry.msgstr, log_detail)
                        
                        for item in detected:
                            detected_items.append({
                                'file_type': 'po',
                                'file_path': po_path,
                                'entry_id': entry.msgid,
                                'entry_context': entry.msgctxt or "",
                                'original_text': entry.msgstr,
                                'sensitive_word': item['keyword'],
                                'category': item['category'],
                                'replacements': item['replacements'],
                                'line_number': entry.linenum if hasattr(entry, 'linenum') else 0,
                                'match_positions': (item['start_pos'], item['end_pos'])
                            })
                
                except Exception as e:
                    print(f"   ⚠️  讀取 PO 檔案失敗：{e}")
        
        # 檢測 JSON 檔案
        if 'json_file' in language_files:
            json_path = language_files['json_file']
            if json_path.exists():
                try:
                    with open(json_path, 'r', encoding='utf-8') as f:
                        json_data = json.load(f)
                    
                    def check_json_recursive(obj, path=""):
                        """遞歸檢查 JSON 物件中的敏感詞"""
                        if isinstance(obj, dict):
                            for key, value in obj.items():
                                new_path = f"{path}.{key}" if path else key
                                check_json_recursive(value, new_path)
                        elif isinstance(obj, list):
                            for i, item in enumerate(obj):
                                new_path = f"{path}[{i}]"
                                check_json_recursive(item, new_path)
                        elif isinstance(obj, str):
                            # 使用優先順序檢測
                            detected = detector.detect_with_priority(obj, log_detail)
                            
                            for item in detected:
                                detected_items.append({
                                    'file_type': 'json',
                                    'file_path': json_path,
                                    'entry_id': path,
                                    'entry_context': "",
                                    'original_text': obj,
                                    'sensitive_word': item['keyword'],
                                    'category': item['category'],
                                    'replacements': item['replacements'],
                                    'line_number': 0,
                                    'match_positions': (item['start_pos'], item['end_pos'])
                                })
                    
                    check_json_recursive(json_data)
                
                except Exception as e:
                    print(f"   ⚠️  讀取 JSON 檔案失敗：{e}")
        
        # 簡化統計輸出
        category_stats = defaultdict(int)
        for item in detected_items:
            category_stats[item['category']] += 1
        
        if detected_items:
            print(f"   📊 檢測到 {len(detected_items)} 個敏感詞")
            for category, count in category_stats.items():
                print(f"     {category}: {count} 個")
        else:
            print(f"   ✅ 無敏感詞")
    
    except Exception as e:
        print(f"   ❌ 檢測錯誤：{e}")
    
    return detected_items


def generate_tobemodified_excel(config, language: str, detected_items: list, output_dir: Path):
    """
    生成待修正 Excel 檔案
    
    Args:
        config: 配置物件
        language: 語言代碼
        detected_items: 檢測到的敏感詞項目列表
        output_dir: 輸出目錄
    """
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    if not detected_items:
        print(f"   ✅ 無需修正項目")
        return
    
    # 建立輸出檔案路徑
    output_file = output_dir / f"{language}_tobemodified.xlsx"
    
    # 創建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{language}_待修正清單"
    
    # 樣式設定
    header_font = Font(bold=True, color="FFFFFF", size=12)
    data_font = Font(size=10)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 取得業態類型
    business_types = config.get_business_types()
    
    # 檢查是否需要添加匹配位置欄位
    keyword_config = config.get_keyword_detection_config()
    inclusion_config = keyword_config.get('inclusion_handling', {})
    add_position_column = inclusion_config.get('add_position_column', False)
    
    # 定義標題列
    headers = [
        "檔案類型", "檔案路徑", "項目ID", "項目內容", "敏感詞", "敏感詞分類"
    ]
    
    # 可選添加匹配位置欄位
    if add_position_column:
        headers.append("匹配位置")
    
    # 為每個業態添加替換方案列和替換結果列
    for bt_code, bt_config in business_types.items():
        headers.append(f"{bt_config['display_name']}_替換方案")
        headers.append(f"{bt_config['display_name']}_替換結果")
    
    # 寫入標題列
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # 寫入數據
    for row_num, item in enumerate(detected_items, 2):
        col_num = 1
        
        # 基本資訊
        basic_data = [
            item['file_type'].upper(),
            str(item['file_path'].name),
            item['entry_id'],
            item['original_text'][:100] + "..." if len(item['original_text']) > 100 else item['original_text'],
            item['sensitive_word'],
            item['category']
        ]
        
        # 可選添加匹配位置
        if add_position_column:
            match_pos = f"{item['match_positions'][0]}-{item['match_positions'][1]}" if 'match_positions' in item else ""
            basic_data.append(match_pos)
        
        for data in basic_data:
            cell = ws.cell(row=row_num, column=col_num, value=data)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            if row_num % 2 == 0:
                cell.fill = alt_row_fill
            
            col_num += 1
        
        # 各業態替換方案和替換結果
        for bt_code, bt_config in business_types.items():
            # 替換方案列
            replacement = item['replacements'].get(bt_code, "")
            cell = ws.cell(row=row_num, column=col_num, value=replacement)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            if row_num % 2 == 0:
                cell.fill = alt_row_fill
            
            col_num += 1
            
            # 替換結果列
            sensitive_word = item['sensitive_word']
            original_text = item['original_text']
            result_value = ""
            
            if replacement and replacement.strip():
                # 使用精確位置替換，而不是簡單的 replace
                if 'match_positions' in item:
                    start_pos, end_pos = item['match_positions']
                    predicted_result = original_text[:start_pos] + replacement + original_text[end_pos:]
                else:
                    # 後備方案：使用普通替換
                    predicted_result = original_text.replace(sensitive_word, replacement)
                
                result_value = predicted_result
            
            cell = ws.cell(row=row_num, column=col_num, value=result_value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            if result_value:
                edit_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                cell.fill = edit_fill
            elif row_num % 2 == 0:
                cell.fill = alt_row_fill
            
            col_num += 1
    
    # 自動調整列寬
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row_idx in range(1, min(ws.max_row + 1, 100)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 確保輸出目錄存在
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 保存檔案
    wb.save(output_file)
    
    print(f"   📄 已生成：{output_file.name} ({len(detected_items)} 個項目)")
    
    # 自動調整列寬
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row_idx in range(1, min(ws.max_row + 1, 100)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 確保輸出目錄存在
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 保存檔案
    wb.save(output_file)
    
    print(f"   📄 已生成：{output_file.name} ({len(detected_items)} 個項目)")


def main():
    """主執行函數 - 簡化輸出版本"""
    print("🚀 開始生成各語言 tobemodified 檔案 (包容關係處理)")
    
    # 載入配置
    config = get_config()
    
    # 檢測可用語言
    available_languages = config.detect_available_languages()
    
    # 取得統一對照表路徑
    excel_path = config.get_comparison_excel_path()
    
    if not excel_path.exists():
        print(f"❌ 找不到對照表檔案：{excel_path}")
        print("   請先執行 generate_phrase_comparison.py 生成對照表")
        return
    
    # 解析語言獨立橫向分區塊 Excel
    try:
        language_blocks = parse_language_blocks_from_excel(excel_path, config)
    except Exception as e:
        print(f"❌ 解析 Excel 檔案失敗：{e}")
        return
    
    if not language_blocks:
        print("❌ Excel 中沒有找到有效的語言區塊")
        return
    
    # 檢查語言匹配
    excel_languages = set(language_blocks.keys())
    input_languages = set(available_languages)
    
    common_languages = excel_languages & input_languages
    
    if not common_languages:
        print("❌ 沒有語言同時存在於語言獨立 Excel 和輸入檔案中")
        print(f"   Excel 中的語言：{list(excel_languages)}")
        print(f"   輸入檔案語言：{list(input_languages)}")
        return
    
    print(f"✅ 將處理 {len(common_languages)} 個語言：{', '.join(sorted(common_languages))}")
    
    # 建立輸出目錄
    try:
        if hasattr(config, 'get_output_dir'):
            output_dir = config.get_output_dir()
        elif hasattr(config, 'output_dir'):
            output_dir = config.output_dir
        elif hasattr(config, 'get_config'):
            config_data = config.get_config()
            output_dir = Path(config_data.get('output_dir', 'i18n_output'))
        else:
            output_dir = Path('i18n_output')
    except Exception as e:
        output_dir = Path('i18n_output')
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 處理每個語言
    total_detected = 0
    processed_languages = 0
    
    for language in sorted(common_languages):
        print(f"\n📋 處理語言：{language}")
        
        sensitive_words = language_blocks[language]
        
        # 使用新的優先順序檢測邏輯
        detected_items = detect_sensitive_phrases_in_files_with_priority(config, language, sensitive_words)
        total_detected += len(detected_items)
        
        # 生成待修正檔案
        generate_tobemodified_excel(config, language, detected_items, output_dir)
        processed_languages += 1
    
    # 生成總結報告
    print(f"\n📊 處理完成：")
    print(f"   處理語言：{processed_languages} 個")
    print(f"   檢測項目：{total_detected} 個")
    print(f"   輸出目錄：{output_dir}")
    
    if total_detected > 0:
        print(f"\n✅ 已生成待修正清單，請檢查並編輯後執行 script_02_apply_fixes.py")
    else:
        print("✅ 所有語言都沒有檢測到敏感詞")


if __name__ == "__main__":
    main()