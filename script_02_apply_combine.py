#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
script_02_apply_combine.py (v1.3 - 修正業態衝突邏輯版)

修正內容：
1. ✅ 修正業態間重複處理同一檔案的問題
2. ✅ 修正衝突檢測邏輯：只處理當前業態的更新
3. ✅ 避免業態間互相干擾
4. ✅ 正確區分真正衝突和正常更新
5. ✅ 改善合併流程邏輯

功能：
1. 選擇要合併的 tobemodified Excel 檔案（支援多選）
2. 選擇 i18n_combine 目錄下的 JSON/PO 檔案作為合併目標
3. 按業態分別處理，避免相互衝突
4. 生成合併後的檔案到 i18n_output/multi_{timestamp}_combined/
5. 提供詳細的合併報告和日誌
"""

import json
import sys
import shutil
import datetime
import argparse
import glob
import re
from pathlib import Path
from collections import defaultdict
from config_loader import get_config

try:
    import openpyxl
    import polib
except ImportError as e:
    print(f"❌ 缺少必要套件：{e}")
    print("請執行：pip install openpyxl polib")
    sys.exit(1)


def detect_tobemodified_files(config) -> dict:
    """檢測可用的 tobemodified 檔案"""
    available_files = {}
    
    # 檢測輸出目錄中的檔案
    try:
        dirs = config.get_directories()
        output_dir = Path(dirs['output_dir'])
    except Exception:
        output_dir = Path('i18n_output')
    
    # 使用配置載入器的語言檢測
    try:
        available_languages = config.detect_available_languages()
    except Exception as e:
        print(f"⚠️  語言檢測失敗：{e}")
        available_languages = []
    
    # 檢測標準命名的檔案
    for language in available_languages:
        tobemodified_path = output_dir / f"{language}_tobemodified.xlsx"
        if tobemodified_path.exists():
            available_files[language] = tobemodified_path
    
    return available_files


def scan_combine_directory(combine_dir: Path) -> dict:
    """掃描 i18n_combine 目錄中的檔案"""
    files = {
        'json': [],
        'po': []
    }
    
    if not combine_dir.exists():
        return files
    
    # 遞歸掃描所有 JSON 和 PO 檔案
    for file_path in combine_dir.rglob("*.json"):
        relative_path = file_path.relative_to(combine_dir)
        files['json'].append({
            'path': file_path,
            'relative_path': str(relative_path),
            'name': file_path.name
        })
    
    for file_path in combine_dir.rglob("*.po"):
        relative_path = file_path.relative_to(combine_dir)
        files['po'].append({
            'path': file_path,
            'relative_path': str(relative_path),
            'name': file_path.name
        })
    
    return files


def choose_tobemodified_files(available_files: dict) -> dict:
    """選擇要使用的 tobemodified 檔案（支援多選）"""
    if not available_files:
        print("❌ 未找到任何 tobemodified 檔案")
        return {}
    
    print("\n📄 可用的 tobemodified 檔案：")
    choices = list(available_files.items())
    
    for i, (language, file_path) in enumerate(choices, 1):
        print(f"  {i}) {language} ({file_path.name})")
    
    print(f"  A) 全部選擇")
    print(f"  0) 取消操作")
    
    selected_files = {}
    
    while True:
        try:
            choice = input(f"\n請選擇要使用的檔案 (可多選，用逗號分隔，如 1,2,3 或 A)：").strip()
            
            if choice == '0':
                print("❌ 操作取消")
                return {}
            elif choice.upper() == 'A':
                selected_files = available_files.copy()
                break
            else:
                # 解析多選
                choice_indices = [int(x.strip()) - 1 for x in choice.split(',')]
                selected_files = {}
                
                for choice_idx in choice_indices:
                    if 0 <= choice_idx < len(choices):
                        language, file_path = choices[choice_idx]
                        selected_files[language] = file_path
                    else:
                        print(f"⚠️  無效選項：{choice_idx + 1}")
                        continue
                
                if selected_files:
                    break
                else:
                    print(f"⚠️  請輸入有效的選項")
                    
        except (ValueError, KeyboardInterrupt):
            print("\n❌ 操作取消")
            return {}
    
    print(f"✅ 選擇了 {len(selected_files)} 個檔案：")
    for language, file_path in selected_files.items():
        print(f"   {language}: {file_path.name}")
    
    return selected_files


def choose_combine_file(files: list, file_type: str) -> Path:
    """選擇要合併的檔案"""
    if not files:
        print(f"⚠️  /i18n_combine/ 中沒有找到 {file_type.upper()} 檔案")
        return None
    
    print(f"\n📁 可用的 {file_type.upper()} 檔案：")
    for i, file_info in enumerate(files, 1):
        print(f"  {i}) {file_info['relative_path']}")
    
    print(f"  0) 跳過 {file_type.upper()} 檔案")
    
    while True:
        try:
            choice = input(f"\n請選擇要合併的 {file_type.upper()} 檔案 (0-{len(files)})：").strip()
            choice_idx = int(choice)
            
            if choice_idx == 0:
                print(f"⏭️  跳過 {file_type.upper()} 檔案")
                return None
            elif 1 <= choice_idx <= len(files):
                selected_file = files[choice_idx - 1]
                print(f"✅ 選擇了：{selected_file['relative_path']}")
                return selected_file['path']
            else:
                print(f"⚠️  請輸入 0-{len(files)} 之間的數字")
        except (ValueError, KeyboardInterrupt):
            print("\n❌ 操作取消")
            return None


def read_excel_updates_for_language(xlsx_path: Path, language: str, config) -> dict:
    """讀取單個語言的 Excel 檔案中的更新資料"""
    try:
        print(f"📖 讀取 {language} 的 Excel 檔案：{xlsx_path.name}")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        header_row = list(ws[1])
        header = {cell.value: idx for idx, cell in enumerate(header_row) if cell.value}
        
        # 基本欄位檢查
        required_columns = ["檔案類型", "項目ID", "項目內容"]
        missing_columns = []
        
        for col in required_columns:
            if col not in header:
                missing_columns.append(col)
        
        if missing_columns:
            print(f"❌ {language} Excel 缺少必要欄位：{missing_columns}")
            return {}
        
        # 自動檢測所有業態的替換結果欄位
        business_types = config.get_business_types()
        available_business_types = []
        
        for bt_code, bt_config in business_types.items():
            display_name = bt_config['display_name']
            result_col_name = f"{display_name}_替換結果"
            if result_col_name in header:
                available_business_types.append(bt_code)
        
        if not available_business_types:
            print(f"❌ {language} 未找到任何業態的替換結果欄位")
            return {}
        
        print(f"   📋 {language} 檢測到業態：{', '.join([business_types[bt]['display_name'] for bt in available_business_types])}")
        
        # 解析更新資料
        updates = {bt_code: {"po": [], "json": []} for bt_code in available_business_types}
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= max(header.values()):
                continue
            
            try:
                file_type = row[header["檔案類型"]]
                entry_id = row[header["項目ID"]]
                original_text = row[header["項目內容"]]
                
                if not file_type or not entry_id:
                    continue
                
                file_type = str(file_type).lower()
                
                # 處理每個可用的業態
                for bt_code in available_business_types:
                    display_name = business_types[bt_code]['display_name']
                    result_col_name = f"{display_name}_替換結果"
                    
                    new_value = row[header[result_col_name]]
                    
                    # 跳過空值和與原文相同的值
                    if not new_value or not str(new_value).strip():
                        continue
                    
                    new_value = str(new_value).strip()
                    
                    if original_text and str(original_text).strip() == new_value:
                        continue
                    
                    # 創建更新記錄，包含語言信息
                    update_record = (str(entry_id), new_value, language)
                    
                    if file_type == "po":
                        updates[bt_code]["po"].append(update_record)
                    elif file_type == "json":
                        updates[bt_code]["json"].append(update_record)
            
            except Exception as e:
                print(f"⚠️  {language} 第 {row_num} 行處理失敗: {e}")
                continue
        
        # 統計有效更新
        total_updates = 0
        for bt_code in available_business_types:
            bt_updates = len(updates[bt_code]["po"]) + len(updates[bt_code]["json"])
            total_updates += bt_updates
            if bt_updates > 0:
                print(f"     {business_types[bt_code]['display_name']}: {bt_updates} 個更新")
        
        print(f"   📊 {language} 總計：{total_updates} 個有效更新")
        return updates
        
    except Exception as e:
        print(f"❌ 讀取 {language} Excel 檔案失敗：{e}")
        return {}


def combine_multilang_json_files_for_business_type(all_updates: dict, target_json_path: Path, 
                                                  output_json_path: Path, bt_code: str, log_detail=None) -> dict:
    """【修正版】為特定業態合併多語言 JSON 檔案，避免業態間衝突，並正確處理數值差異"""
    result = {
        "success": False,
        "merged": 0,
        "skipped": 0,
        "conflicts": [],
        "errors": [],
        "language_stats": {}
    }
    
    # 檢查是否有當前業態的更新
    has_updates = False
    for language_updates in all_updates.values():
        if bt_code in language_updates and language_updates[bt_code]['json']:
            has_updates = True
            break
    
    if not has_updates:
        result["success"] = True
        if log_detail:
            log_detail(f"JSON ({bt_code}): 沒有任何更新項目")
        return result
    
    try:
        # 載入目標 JSON 檔案
        if not target_json_path.exists():
            result["errors"].append(f"目標 JSON 檔案不存在：{target_json_path}")
            return result
        
        target_data = json.loads(target_json_path.read_text(encoding="utf-8"))
        print(f"   📄 載入目標多語言 JSON 檔案：{target_json_path.name}")
        if log_detail:
            log_detail(f"載入目標 JSON 檔案：{target_json_path.name}")
        
        # 檢查是否為多語言結構
        is_multilang_structure = check_multilang_json_structure(target_data)
        print(f"   🔍 多語言結構檢測：{'是' if is_multilang_structure else '否'}")
        if log_detail:
            log_detail(f"多語言結構檢測：{'是' if is_multilang_structure else '否'}")
        
        conflicts = []
        language_stats = {}
        
        # 只處理當前業態的更新
        for language, language_updates in all_updates.items():
            if bt_code not in language_updates:
                continue
                
            language_stats[language] = {"merged": 0, "skipped": 0, "conflicts": 0}
            
            if log_detail:
                log_detail(f"處理語言 {language} 的 JSON 更新 (業態: {bt_code})")
            
            # 處理當前業態的 JSON 更新
            bt_updates = language_updates[bt_code]
            for json_path_str, new_value, update_language in bt_updates['json']:
                if log_detail:
                    log_detail(f"處理更新：{update_language}.{json_path_str} = {new_value}")
                
                # 多語言結構的路徑映射
                if is_multilang_structure:
                    multilang_path = f"{update_language}.{json_path_str}"
                else:
                    multilang_path = json_path_str
                
                # 獲取現有值
                existing_value = get_json_value_by_path(target_data, multilang_path)
                
                # 【修正關鍵邏輯】正確處理值的比較和衝突檢測
                if existing_value is not None:
                    existing_str = str(existing_value).strip()
                    new_str = str(new_value).strip()
                    
                    # 如果值完全相同，跳過
                    if existing_str == new_str:
                        result["skipped"] += 1
                        language_stats[update_language]["skipped"] += 1
                        if log_detail:
                            log_detail(f"跳過相同值：{multilang_path} = '{new_str}'")
                        continue
                    
                    # 【重要修正】當值不同時，應該標記為衝突並讓用戶決定
                    if existing_str != new_str:
                        conflict_info = {
                            "path": multilang_path,
                            "language": update_language,
                            "existing_value": existing_str,
                            "new_value": new_str,
                            "file_type": "json"
                        }
                        conflicts.append(conflict_info)
                        result["conflicts"].append(conflict_info)
                        language_stats[update_language]["conflicts"] += 1
                        
                        if log_detail:
                            log_detail(f"發現衝突：{multilang_path}")
                            log_detail(f"  現有值: '{existing_str}'")
                            log_detail(f"  新值: '{new_str}'")
                        
                        # 詢問用戶如何處理衝突
                        choice = handle_json_conflict(multilang_path, existing_str, new_str, update_language)
                        
                        if choice == "keep_existing":
                            result["skipped"] += 1
                            language_stats[update_language]["skipped"] += 1
                            if log_detail:
                                log_detail(f"保留現有值：{multilang_path} = '{existing_str}'")
                            continue
                        elif choice == "use_new":
                            # 繼續執行更新邏輯
                            if log_detail:
                                log_detail(f"採用新值：{multilang_path} = '{new_str}'")
                        elif choice == "skip":
                            result["skipped"] += 1
                            language_stats[update_language]["skipped"] += 1
                            if log_detail:
                                log_detail(f"跳過處理：{multilang_path}")
                            continue
                
                # 應用更新
                if set_json_value_by_path(target_data, multilang_path, new_value):
                    result["merged"] += 1
                    language_stats[update_language]["merged"] += 1
                    if log_detail:
                        original_display = f"'{existing_value}'" if existing_value is not None else "無"
                        log_detail(f"成功更新：{multilang_path} = '{new_value}' (原值: {original_display})")
                else:
                    error_msg = f"無法設置 JSON 路徑：{multilang_path} (語言: {update_language})"
                    result["errors"].append(error_msg)
                    if log_detail:
                        log_detail(f"錯誤：{error_msg}")
        
        # 保存合併後的檔案
        output_json_path.parent.mkdir(parents=True, exist_ok=True)
        
        json_content = json.dumps(target_data, ensure_ascii=False, indent=2)
        output_json_path.write_text(json_content, encoding="utf-8")
        
        result["success"] = True
        result["language_stats"] = language_stats
        
        # 修正日誌訊息，包含衝突數量
        total_conflicts = len(conflicts)
        if log_detail:
            log_detail(f"JSON ({bt_code}) 合併完成：合併 {result['merged']} 個，跳過 {result['skipped']} 個，衝突 {total_conflicts} 個")
        
    except json.JSONDecodeError as e:
        error_msg = f"JSON 格式錯誤：{e}"
        result["errors"].append(error_msg)
        if log_detail:
            log_detail(f"錯誤：{error_msg}")
    except Exception as e:
        error_msg = f"JSON 檔案合併失敗：{e}"
        result["errors"].append(error_msg)
        if log_detail:
            log_detail(f"錯誤：{error_msg}")
    
    return result


def handle_json_conflict(path: str, existing_value: str, new_value: str, language: str) -> str:
    """處理 JSON 合併衝突，讓用戶選擇如何處理"""
    print(f"\n⚠️  發現衝突：")
    print(f"📍 路徑：{path}")
    print(f"🌍 語言：{language}")
    print(f"📄 現有值：'{existing_value}'")
    print(f"🆕 新值：'{new_value}'")
    
    while True:
        print(f"\n請選擇處理方式：")
        print(f"  1) 保留現有值 ('{existing_value}')")
        print(f"  2) 使用新值 ('{new_value}')")
        print(f"  3) 跳過此項目")
        print(f"  A) 對所有類似衝突使用新值")
        print(f"  K) 對所有類似衝突保留現有值")
        
        try:
            choice = input(f"請選擇 (1/2/3/A/K)：").strip().upper()
            
            if choice == "1":
                return "keep_existing"
            elif choice == "2":
                return "use_new"
            elif choice == "3":
                return "skip"
            elif choice == "A":
                # 可以擴展為全局策略
                print(f"✅ 將使用新值")
                return "use_new"
            elif choice == "K":
                # 可以擴展為全局策略
                print(f"✅ 將保留現有值")
                return "keep_existing"
            else:
                print(f"⚠️  請輸入有效選項 (1/2/3/A/K)")
                
        except KeyboardInterrupt:
            print(f"\n❌ 操作取消，跳過此項目")
            return "skip"


def generate_conflict_report(conflicts: list, output_dir: Path, timestamp: str):
    """生成衝突報告"""
    if not conflicts:
        return
    
    conflict_report_file = output_dir / f"conflicts_report_{timestamp}.txt"
    
    try:
        with open(conflict_report_file, 'w', encoding='utf-8') as f:
            f.write(f"JSON 合併衝突報告\n")
            f.write(f"生成時間：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"{'='*60}\n\n")
            
            f.write(f"總衝突數量：{len(conflicts)}\n\n")
            
            for i, conflict in enumerate(conflicts, 1):
                f.write(f"衝突 {i}：\n")
                f.write(f"  路徑：{conflict['path']}\n")
                f.write(f"  語言：{conflict['language']}\n")
                f.write(f"  現有值：'{conflict['existing_value']}'\n")
                f.write(f"  新值：'{conflict['new_value']}'\n")
                f.write(f"  檔案類型：{conflict['file_type']}\n")
                f.write(f"\n{'-'*40}\n\n")
            
            f.write(f"處理建議：\n")
            f.write(f"1. 檢查值的差異是否為預期的更新\n")
            f.write(f"2. 確認語言翻譯的正確性\n")
            f.write(f"3. 驗證業態特定的術語使用\n")
            f.write(f"4. 考慮建立翻譯一致性檢查機制\n")
        
        print(f"📄 衝突報告已生成：{conflict_report_file}")
        
    except Exception as e:
        print(f"⚠️  生成衝突報告失敗：{e}")

def check_multilang_json_structure(data: dict) -> bool:
    """檢查 JSON 是否為多語言結構"""
    if not isinstance(data, dict):
        return False
    
    # 檢查頂層 key 是否像語言代碼
    for key in data.keys():
        if isinstance(key, str) and re.match(r'^[a-z]{2}(-[A-Z]{2})?$', key):
            # 如果至少有一個 key 像語言代碼，且其值是字典，則認為是多語言結構
            if isinstance(data[key], dict):
                return True
    
    return False


def combine_po_files_for_business_type(all_updates: dict, target_po_path: Path, 
                                     output_dir: Path, bt_code: str, log_detail=None) -> dict:
    """【修正版】為特定業態處理 PO 檔案合併"""
    result = {
        "success": False,
        "merged": 0,
        "skipped": 0,
        "conflicts": [],
        "errors": [],
        "language_stats": {}
    }
    
    # 檢查是否有當前業態的 PO 更新
    has_updates = False
    for language_updates in all_updates.values():
        if bt_code in language_updates and language_updates[bt_code]['po']:
            has_updates = True
            break
    
    if not has_updates:
        result["success"] = True
        if log_detail:
            log_detail(f"PO ({bt_code}): 沒有任何更新項目")
        return result
    
    try:
        # 載入目標 PO 檔案
        if not target_po_path.exists():
            result["errors"].append(f"目標 PO 檔案不存在：{target_po_path}")
            return result
        
        target_po = polib.pofile(str(target_po_path))
        print(f"   📄 載入目標 PO 檔案：{target_po_path.name}，共 {len(target_po)} 個條目")
        if log_detail:
            log_detail(f"載入目標 PO 檔案：{target_po_path.name}，共 {len(target_po)} 個條目")
        
        language_stats = {}
        
        # 【修正】只處理當前業態的更新
        for language, language_updates in all_updates.items():
            if bt_code not in language_updates:
                continue
                
            language_stats[language] = {"merged": 0, "skipped": 0, "conflicts": 0}
            
            # 處理當前業態的 PO 更新
            bt_updates = language_updates[bt_code]
            for msgid, new_msgstr, update_language in bt_updates['po']:
                target_entry = target_po.find(msgid)
                
                if target_entry:
                    # 【修正】只有當現有值和新值真的不同時才需要更新
                    if target_entry.msgstr and target_entry.msgstr.strip():
                        if target_entry.msgstr == new_msgstr:
                            # 值相同，跳過
                            result["skipped"] += 1
                            language_stats[update_language]["skipped"] += 1
                            continue
                    
                    # 應用更新
                    target_entry.msgstr = new_msgstr
                    result["merged"] += 1
                    language_stats[update_language]["merged"] += 1
                else:
                    # 目標檔案中沒有此條目，添加新條目
                    new_entry = polib.POEntry(
                        msgid=msgid,
                        msgstr=new_msgstr
                    )
                    target_po.append(new_entry)
                    result["merged"] += 1
                    language_stats[update_language]["merged"] += 1
        
        # 保存合併後的檔案
        config = get_config()
        business_types = config.get_business_types()
        
        if bt_code in business_types:
            suffix = business_types[bt_code]['suffix']
            output_po_path = output_dir / f"{target_po_path.stem}{suffix}_combined.po"
            output_po_path.parent.mkdir(parents=True, exist_ok=True)
            target_po.save(str(output_po_path))
        
        result["success"] = True
        result["language_stats"] = language_stats
        
        if log_detail:
            log_detail(f"PO ({bt_code}) 合併完成：合併 {result['merged']} 個，跳過 {result['skipped']} 個")
        
    except Exception as e:
        error_msg = f"PO 檔案合併失敗：{e}"
        result["errors"].append(error_msg)
        if log_detail:
            log_detail(f"錯誤：{error_msg}")
    
    return result


def get_json_value_by_path(data: dict, path: str):
    """按路徑獲取 JSON 值"""
    try:
        path_parts = parse_json_path(path)
        current = data
        
        for part_type, part_value in path_parts:
            if part_type == 'key':
                if part_value not in current:
                    return None
                current = current[part_value]
            elif part_type == 'index':
                if not isinstance(current, list) or len(current) <= part_value:
                    return None
                current = current[part_value]
        
        return current
        
    except Exception:
        return None


def set_json_value_by_path(data: dict, path: str, new_value: str) -> bool:
    """按路徑設置 JSON 值"""
    try:
        path_parts = parse_json_path(path)
        current = data
        
        for i, (part_type, part_value) in enumerate(path_parts):
            is_last = (i == len(path_parts) - 1)
            
            if part_type == 'key':
                if is_last:
                    current[part_value] = new_value
                else:
                    if part_value not in current:
                        next_part_type = path_parts[i + 1][0] if i + 1 < len(path_parts) else 'key'
                        current[part_value] = [] if next_part_type == 'index' else {}
                    current = current[part_value]
            
            elif part_type == 'index':
                if is_last:
                    while len(current) <= part_value:
                        current.append(None)
                    current[part_value] = new_value
                else:
                    while len(current) <= part_value:
                        current.append(None)
                    if current[part_value] is None:
                        next_part_type = path_parts[i + 1][0] if i + 1 < len(path_parts) else 'key'
                        current[part_value] = [] if next_part_type == 'index' else {}
                    current = current[part_value]
        
        return True
        
    except Exception as e:
        return False


def parse_json_path(path: str) -> list:
    """解析 JSON 路徑"""
    parts = []
    current = ""
    in_bracket = False
    
    for char in path:
        if char == '[':
            if current:
                parts.append(('key', current))
                current = ""
            in_bracket = True
        elif char == ']':
            if in_bracket and current:
                try:
                    parts.append(('index', int(current)))
                except ValueError:
                    raise ValueError(f"無效的陣列索引：{current}")
                current = ""
            in_bracket = False
        elif char == '.' and not in_bracket:
            if current:
                parts.append(('key', current))
                current = ""
        else:
            current += char
    
    if current:
        parts.append(('key', current))
    
    return parts


def main():
    """主執行函數"""
    print("🚀 開始多語言檔案合併處理 (v1.3 - 修正業態衝突邏輯版)")
    
    # 載入配置
    config = get_config()
    
    # 檢測可用的 tobemodified 檔案
    available_files = detect_tobemodified_files(config)
    
    if not available_files:
        print("❌ 未找到任何 tobemodified 檔案")
        print("請先執行 script_01_generate_xlsx.py 生成檔案")
        sys.exit(1)
    
    # 步驟1：選擇 tobemodified 檔案（支援多選）
    selected_files = choose_tobemodified_files(available_files)
    if not selected_files:
        sys.exit(1)
    
    # 檢查 i18n_combine 目錄
    combine_dir = Path("i18n_combine")
    
    if not combine_dir.exists():
        print(f"❌ 合併目錄不存在：{combine_dir}")
        print(f"請創建目錄並放入要合併的檔案")
        sys.exit(1)
    
    print(f"📁 掃描合併目錄：{combine_dir}")
    
    # 掃描 combine 目錄中的檔案
    combine_files = scan_combine_directory(combine_dir)
    
    # 步驟2：選擇要合併的 JSON 檔案
    target_json_path = choose_combine_file(combine_files['json'], 'json')
    
    # 步驟3：選擇要合併的 PO 檔案
    target_po_path = choose_combine_file(combine_files['po'], 'po')
    
    # 檢查是否至少選擇了一個檔案
    if not target_json_path and not target_po_path:
        print("❌ 必須至少選擇一個檔案進行合併")
        sys.exit(1)
    
    # 讀取所有選中語言的 Excel 更新資料
    all_updates = {}
    for language, xlsx_path in selected_files.items():
        updates = read_excel_updates_for_language(xlsx_path, language, config)
        if updates:
            all_updates[language] = updates
    
    if not all_updates:
        print("❌ 沒有讀取到任何有效的更新資料")
        sys.exit(1)
    
    # 統計所有業態
    all_business_types = set()
    for language_updates in all_updates.values():
        all_business_types.update(language_updates.keys())
    
    print(f"\n📋 合併設定：")
    print(f"   來源語言：{', '.join(selected_files.keys())}")
    if target_json_path:
        print(f"   JSON 檔案：{target_json_path.relative_to(combine_dir)}")
    if target_po_path:
        print(f"   PO 檔案：{target_po_path.relative_to(combine_dir)}")
    print(f"   涵蓋業態：{', '.join([config.get_business_types()[bt]['display_name'] for bt in all_business_types])}")
    
    # 建立輸出目錄 - 使用正確的命名格式
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    dirs = config.get_directories()
    output_dir = Path(dirs['output_dir']) / f"multi_{timestamp}_combined"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 設置日誌
    log_file = output_dir / f"multi_combine_{timestamp}.log"
    
    def log_detail(message: str):
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(f"{datetime.datetime.now().strftime('%H:%M:%S')} - {message}\n")
    
    log_detail(f"開始多語言合併處理")
    log_detail(f"語言：{', '.join(selected_files.keys())}")
    log_detail(f"來源檔案：{list(selected_files.values())}")
    log_detail(f"涵蓋業態：{', '.join(all_business_types)}")
    
    # 【修正】處理合併邏輯 - 避免業態間衝突
    business_types = config.get_business_types()
    all_results = {}
    
    # 【修正】按業態分別處理，避免相互干擾
    for bt_code in all_business_types:
        if bt_code not in business_types:
            continue
            
        bt_config = business_types[bt_code]
        display_name = bt_config['display_name']
        suffix = bt_config['suffix']
        
        print(f"\n📝 處理 {display_name}...")
        log_detail(f"開始處理業態：{display_name}")
        
        results = {}
        
        # 【修正】為當前業態處理 JSON 檔案
        if target_json_path:
            output_json_path = output_dir / f"{target_json_path.stem}{suffix}_combined.json"
            json_result = combine_multilang_json_files_for_business_type(
                all_updates,
                target_json_path,
                output_json_path,
                bt_code,
                log_detail
            )
            results['json_result'] = json_result
            
            # 顯示結果
            if json_result.get('errors'):
                print(f"     ❌ JSON 檔案處理錯誤：{json_result['errors']}")
            else:
                # 顯示語言統計
                if json_result.get('language_stats'):
                    for lang, stats in json_result['language_stats'].items():
                        if stats['merged'] > 0 or stats['skipped'] > 0:
                            print(f"     📊 {lang}: 合併 {stats['merged']} 個, 跳過 {stats['skipped']} 個")
                
                if json_result.get('merged', 0) == 0 and json_result.get('skipped', 0) == 0:
                    print(f"     ℹ️  {display_name} 沒有 JSON 更新項目")
        
        # 【修正】為當前業態處理 PO 檔案
        if target_po_path:
            po_result = combine_po_files_for_business_type(
                all_updates,
                target_po_path,
                output_dir,
                bt_code,
                log_detail
            )
            results['po_result'] = po_result
            
            # 顯示結果
            if po_result.get('errors'):
                print(f"     ❌ PO 檔案處理錯誤：{po_result['errors']}")
            else:
                # 顯示語言統計
                if po_result.get('language_stats'):
                    for lang, stats in po_result['language_stats'].items():
                        if stats['merged'] > 0 or stats['skipped'] > 0:
                            print(f"     📊 {lang}: 合併 {stats['merged']} 個, 跳過 {stats['skipped']} 個")
                
                if po_result.get('merged', 0) == 0 and po_result.get('skipped', 0) == 0:
                    print(f"     ℹ️  {display_name} 沒有 PO 更新項目")
        
        # 如果沒有更新，複製原檔案
        if target_json_path and results.get('json_result', {}).get('merged', 0) == 0:
            output_json_path = output_dir / f"{target_json_path.stem}{suffix}_combined.json"
            if not output_json_path.exists():
                output_json_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(target_json_path, output_json_path)
                print(f"     📄 複製 JSON 檔案（無更新）")
                log_detail(f"複製原始 JSON 檔案：{target_json_path.name}")
        
        if target_po_path and results.get('po_result', {}).get('merged', 0) == 0:
            output_po_path = output_dir / f"{target_po_path.stem}{suffix}_combined.po"
            if not output_po_path.exists():
                output_po_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(target_po_path, output_po_path)
                print(f"     📄 複製 PO 檔案（無更新）")
                log_detail(f"複製原始 PO 檔案：{target_po_path.name}")
        
        all_results[bt_code] = results
        
        # 統計結果
        total_merged = 0
        total_skipped = 0
        total_errors = 0
        
        for result in results.values():
            total_merged += result.get('merged', 0)
            total_skipped += result.get('skipped', 0)
            total_errors += len(result.get('errors', []))
        
        if total_errors > 0:
            print(f"     ❌ 處理失敗 - 錯誤: {total_errors} 個")
        else:
            print(f"     ✅ 完成 - 合併: {total_merged} 個, 跳過: {total_skipped} 個")
        
        log_detail(f"{display_name} 處理完成：合併 {total_merged} 個，跳過 {total_skipped} 個，錯誤 {total_errors} 個")
    
    # 生成最終報告
    total_merged = sum(
        sum(result.get('merged', 0) for result in results.values())
        for results in all_results.values()
    )
    total_skipped = sum(
        sum(result.get('skipped', 0) for result in results.values())
        for results in all_results.values()
    )
    total_errors = sum(
        sum(len(result.get('errors', [])) for result in results.values())
        for results in all_results.values()
    )
    
    print(f"\n🎉 多語言合併處理完成！")
    print(f"📊 處理結果：合併 {total_merged} 個項目，跳過 {total_skipped} 個項目")
    if total_errors > 0:
        print(f"⚠️  處理錯誤：{total_errors} 個")
    print(f"📁 輸出目錄：{output_dir}")
    
    # 生成處理摘要
    generate_multilang_summary_report(all_results, all_updates, output_dir, timestamp, target_json_path, target_po_path, log_detail)


def generate_multilang_summary_report(results: dict, all_updates: dict, output_dir: Path, timestamp: str, 
                                     target_json_path: Path, target_po_path: Path, log_detail):
    """生成多語言合併處理摘要報告"""
    summary_file = output_dir / f"multi_combine_summary_{timestamp}.txt"
    
    try:
        with open(summary_file, 'w', encoding='utf-8') as f:
            f.write(f"多語言檔案合併處理摘要報告\n")
            f.write(f"生成時間：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"{'='*60}\n\n")
            
            f.write(f"目標檔案：\n")
            if target_json_path:
                f.write(f"  JSON: {target_json_path}\n")
            if target_po_path:
                f.write(f"  PO: {target_po_path}\n")
            f.write(f"\n")
            
            f.write(f"處理的語言：\n")
            for language in all_updates.keys():
                f.write(f"  - {language}\n")
            f.write(f"\n")
            
            total_merged = 0
            total_skipped = 0
            total_errors = 0
            successful_business_types = []
            failed_business_types = []
            
            # 按業態統計
            for bt_code, bt_results in results.items():
                f.write(f"業態：{bt_code}\n")
                
                bt_merged = sum(result.get('merged', 0) for result in bt_results.values())
                bt_skipped = sum(result.get('skipped', 0) for result in bt_results.values())
                bt_errors = []
                for result in bt_results.values():
                    bt_errors.extend(result.get('errors', []))
                
                f.write(f"合併數量：{bt_merged}\n")
                f.write(f"跳過數量：{bt_skipped}\n")
                
                # 語言級別統計
                f.write(f"語言統計：\n")
                for result in bt_results.values():
                    if 'language_stats' in result:
                        for lang, stats in result['language_stats'].items():
                            f.write(f"  {lang}: 合併 {stats['merged']}, 跳過 {stats['skipped']}, 衝突 {stats.get('conflicts', 0)}\n")
                
                if bt_errors:
                    f.write(f"錯誤：\n")
                    for error in bt_errors:
                        f.write(f"  - {error}\n")
                    failed_business_types.append(bt_code)
                else:
                    successful_business_types.append(bt_code)
                
                total_merged += bt_merged
                total_skipped += bt_skipped
                total_errors += len(bt_errors)
                
                f.write(f"\n{'-'*40}\n\n")
            
            # 總計統計
            f.write(f"處理總結：\n")
            f.write(f"成功業態：{len(successful_business_types)}\n")
            f.write(f"失敗業態：{len(failed_business_types)}\n")
            f.write(f"總合併項目：{total_merged}\n")
            f.write(f"總跳過項目：{total_skipped}\n")
            f.write(f"總錯誤項目：{total_errors}\n")
            f.write(f"處理語言數：{len(all_updates)}\n")
            
            if successful_business_types:
                f.write(f"\n成功的業態：{', '.join(successful_business_types)}\n")
            
            if failed_business_types:
                f.write(f"失敗的業態：{', '.join(failed_business_types)}\n")
            
            f.write(f"\n多語言合併說明：\n")
            f.write(f"- 本次處理支援多個語言的 tobemodified 合併到同一檔案\n")
            f.write(f"- JSON 檔案支援多語言結構（如 enterprise.json）\n")
            f.write(f"- 自動檢測並處理語言層級的路徑映射\n")
            f.write(f"- 按業態分別處理，避免業態間相互干擾\n")
            f.write(f"- 相同 key 且相同 value 的項目會自動跳過\n")
            f.write(f"- 不同 value 的項目會正常更新（不再視為衝突）\n")
            
            f.write(f"\n使用建議：\n")
            f.write(f"- 確認目標 JSON 檔案採用多語言結構（頂層為語言代碼）\n")
            f.write(f"- 合併前建議備份原始檔案\n")
            f.write(f"- 合併後請測試多語言翻譯檔案的正確性\n")
            f.write(f"- 檢查各語言層級的數據完整性\n")
            
            # 修正版本說明
            f.write(f"\n修正版本 v1.3 改進：\n")
            f.write(f"- 修正業態間重複處理同一檔案的問題\n")
            f.write(f"- 修正衝突檢測邏輯：只處理當前業態的更新\n")
            f.write(f"- 避免業態間互相干擾\n")
            f.write(f"- 正確區分真正衝突和正常更新\n")
            f.write(f"- 改善合併流程邏輯\n")
        
        log_detail(f"多語言合併摘要報告已生成：{summary_file}")
        
    except Exception as e:
        log_detail(f"生成多語言合併摘要報告失敗：{e}")


if __name__ == "__main__":
    main()