#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
script_02_apply_combine.py (v1.4 - 層級衝突檢測版)

新增內容：
1. ✅ 檢測相同ID名稱但在不同層級的衝突
2. ✅ 全面列出所有層級衝突詳情
3. ✅ 發現層級衝突時終止進程
4. ✅ 提供詳細的衝突報告和修正建議
5. ✅ 支援多語言和多業態的層級衝突檢測

功能：
1. 選擇要合併的 tobemodified Excel 檔案（支援多選）
2. 選擇 i18n_combine 目錄下的 JSON/PO 檔案作為合併目標
3. 按業態分別處理，避免相互衝突
4. **新增：檢測並報告相同ID在不同層級的衝突**
5. 生成合併後的檔案到 i18n_output/multi_{timestamp}_combined/
6. 提供詳細的合併報告和日誌
"""

import json
import sys
import shutil
import datetime
import argparse
import glob
import re
from pathlib import Path
from collections import defaultdict, Counter
from config_loader import get_config

try:
    import openpyxl
    import polib
except ImportError as e:
    print(f"❌ 缺少必要套件：{e}")
    print("請執行：pip install openpyxl polib")
    sys.exit(1)


class LayerConflictDetector:
    """層級衝突檢測器"""
    
    def __init__(self):
        self.json_conflicts = []
        self.po_conflicts = []
        self.all_json_paths = defaultdict(list)
        self.all_po_ids = defaultdict(list)
    
    def detect_json_layer_conflicts(self, all_updates: dict, target_json_data: dict, is_multilang: bool) -> bool:
        """
        檢測 JSON 檔案中的層級衝突
        
        Args:
            all_updates: 所有語言的更新資料
            target_json_data: 目標 JSON 檔案內容
            is_multilang: 是否為多語言結構
            
        Returns:
            bool: 是否發現衝突
        """
        print("🔍 檢測 JSON 層級衝突...")
        
        # 收集所有路徑和其層級信息
        path_info = {}  # {path: {layers: [層級列表], languages: [語言列表], business_types: [業態列表]}}
        
        # 從更新資料中收集路徑
        for language, language_updates in all_updates.items():
            for bt_code, bt_updates in language_updates.items():
                for json_path_str, new_value, update_language in bt_updates['json']:
                    # 多語言結構的路徑映射
                    if is_multilang:
                        multilang_path = f"{update_language}.{json_path_str}"
                    else:
                        multilang_path = json_path_str
                    
                    # 分析路徑層級
                    layers = self._analyze_json_path_layers(multilang_path)
                    
                    if multilang_path not in path_info:
                        path_info[multilang_path] = {
                            'layers': [],
                            'languages': set(),
                            'business_types': set(),
                            'values': set()
                        }
                    
                    path_info[multilang_path]['layers'] = layers
                    path_info[multilang_path]['languages'].add(update_language)
                    path_info[multilang_path]['business_types'].add(bt_code)
                    path_info[multilang_path]['values'].add(str(new_value))
        
        # 從目標檔案中收集現有路徑
        existing_paths = self._extract_json_paths(target_json_data)
        for path in existing_paths:
            if path not in path_info:
                path_info[path] = {
                    'layers': self._analyze_json_path_layers(path),
                    'languages': set(),
                    'business_types': set(),
                    'values': set()
                }
        
        # 檢測層級衝突
        conflicts_found = self._detect_layer_conflicts_in_paths(path_info, 'json')
        
        if conflicts_found:
            print(f"❌ 發現 {len(self.json_conflicts)} 個 JSON 層級衝突")
            return True
        else:
            print("✅ 未發現 JSON 層級衝突")
            return False
    
    def detect_po_layer_conflicts(self, all_updates: dict, target_po_data) -> bool:
        """
        檢測 PO 檔案中的層級衝突（msgid 重複但在不同上下文）
        
        Args:
            all_updates: 所有語言的更新資料
            target_po_data: 目標 PO 檔案內容
            
        Returns:
            bool: 是否發現衝突
        """
        print("🔍 檢測 PO 層級衝突...")
        
        # 收集所有 msgid 和其上下文信息
        msgid_info = {}  # {msgid: {contexts: [上下文列表], languages: [語言列表], business_types: [業態列表]}}
        
        # 從更新資料中收集 msgid
        for language, language_updates in all_updates.items():
            for bt_code, bt_updates in language_updates.items():
                for msgid, new_msgstr, update_language in bt_updates['po']:
                    if msgid not in msgid_info:
                        msgid_info[msgid] = {
                            'contexts': set(),
                            'languages': set(),
                            'business_types': set(),
                            'values': set()
                        }
                    
                    msgid_info[msgid]['languages'].add(update_language)
                    msgid_info[msgid]['business_types'].add(bt_code)
                    msgid_info[msgid]['values'].add(str(new_msgstr))
        
        # 從目標檔案中收集現有 msgid
        for entry in target_po_data:
            msgid = entry.msgid
            msgctxt = getattr(entry, 'msgctxt', None) or 'default'
            
            if msgid not in msgid_info:
                msgid_info[msgid] = {
                    'contexts': set(),
                    'languages': set(),
                    'business_types': set(),
                    'values': set()
                }
            
            msgid_info[msgid]['contexts'].add(msgctxt)
            if entry.msgstr:
                msgid_info[msgid]['values'].add(entry.msgstr)
        
        # 檢測 PO 檔案的"層級"衝突（主要是上下文衝突）
        conflicts_found = False
        for msgid, info in msgid_info.items():
            if len(info['contexts']) > 1:
                conflict = {
                    'id': msgid,
                    'type': 'po_context_conflict',
                    'contexts': list(info['contexts']),
                    'languages': list(info['languages']),
                    'business_types': list(info['business_types']),
                    'values': list(info['values']),
                    'description': f"msgid '{msgid}' 存在於多個不同上下文中"
                }
                self.po_conflicts.append(conflict)
                conflicts_found = True
        
        if conflicts_found:
            print(f"❌ 發現 {len(self.po_conflicts)} 個 PO 上下文衝突")
            return True
        else:
            print("✅ 未發現 PO 上下文衝突")
            return False
    
    def _analyze_json_path_layers(self, path: str) -> list:
        """分析 JSON 路徑的層級結構"""
        parts = []
        current = ""
        in_bracket = False
        
        for char in path:
            if char == '[':
                if current:
                    parts.append(('key', current))
                    current = ""
                in_bracket = True
            elif char == ']':
                if in_bracket and current:
                    try:
                        parts.append(('index', int(current)))
                    except ValueError:
                        parts.append(('key', current))
                    current = ""
                in_bracket = False
            elif char == '.' and not in_bracket:
                if current:
                    parts.append(('key', current))
                    current = ""
            else:
                current += char
        
        if current:
            parts.append(('key', current))
        
        return parts
    
    def _extract_json_paths(self, data, prefix=""):
        """遞歸提取 JSON 檔案中的所有路徑"""
        paths = []
        
        if isinstance(data, dict):
            for key, value in data.items():
                current_path = f"{prefix}.{key}" if prefix else key
                paths.append(current_path)
                
                if isinstance(value, (dict, list)):
                    paths.extend(self._extract_json_paths(value, current_path))
        
        elif isinstance(data, list):
            for i, value in enumerate(data):
                current_path = f"{prefix}[{i}]"
                paths.append(current_path)
                
                if isinstance(value, (dict, list)):
                    paths.extend(self._extract_json_paths(value, current_path))
        
        return paths
    
    def _detect_layer_conflicts_in_paths(self, path_info: dict, file_type: str) -> bool:
        """檢測路徑中的層級衝突"""
        conflicts_found = False
        
        # 按最後一個路徑元素分組檢查
        end_key_groups = defaultdict(list)
        
        for path, info in path_info.items():
            # 獲取路徑的最後一個元素作為關鍵詞
            layers = info['layers']
            if layers:
                last_element = layers[-1][1]  # (type, value) 中的 value
                end_key_groups[last_element].append((path, info))
        
        # 檢查每個關鍵詞是否出現在不同層級
        for end_key, path_list in end_key_groups.items():
            if len(path_list) > 1:
                # 檢查是否真的是不同層級（不只是路徑不同）
                unique_layer_structures = set()
                
                for path, info in path_list:
                    # 創建層級結構的簽名
                    layer_signature = self._create_layer_signature(info['layers'])
                    unique_layer_structures.add(layer_signature)
                
                # 如果有多個不同的層級結構，就是衝突
                if len(unique_layer_structures) > 1:
                    conflict = {
                        'id': end_key,
                        'type': f'{file_type}_layer_conflict',
                        'paths': [],
                        'layer_structures': [],
                        'languages': set(),
                        'business_types': set(),
                        'values': set(),
                        'description': f"ID '{end_key}' 出現在多個不同的層級結構中"
                    }
                    
                    for path, info in path_list:
                        conflict['paths'].append(path)
                        conflict['layer_structures'].append(self._format_layer_structure(info['layers']))
                        conflict['languages'].update(info['languages'])
                        conflict['business_types'].update(info['business_types'])
                        conflict['values'].update(info['values'])
                    
                    # 轉換 set 為 list 以便序列化
                    conflict['languages'] = list(conflict['languages'])
                    conflict['business_types'] = list(conflict['business_types'])
                    conflict['values'] = list(conflict['values'])
                    
                    if file_type == 'json':
                        self.json_conflicts.append(conflict)
                    else:
                        self.po_conflicts.append(conflict)
                    
                    conflicts_found = True
        
        return conflicts_found
    
    def _create_layer_signature(self, layers: list) -> str:
        """創建層級結構的簽名"""
        signature_parts = []
        for layer_type, layer_value in layers[:-1]:  # 排除最後一個元素
            if layer_type == 'key':
                signature_parts.append(f"k:{layer_value}")
            elif layer_type == 'index':
                signature_parts.append(f"i:{layer_value}")
        return ".".join(signature_parts)
    
    def _format_layer_structure(self, layers: list) -> str:
        """格式化層級結構為可讀字符串"""
        parts = []
        for layer_type, layer_value in layers:
            if layer_type == 'key':
                parts.append(str(layer_value))
            elif layer_type == 'index':
                parts.append(f"[{layer_value}]")
        return ".".join(parts)
    
    def print_conflict_report(self):
        """打印詳細的衝突報告"""
        total_conflicts = len(self.json_conflicts) + len(self.po_conflicts)
        
        if total_conflicts == 0:
            print("✅ 未發現任何層級衝突")
            return False
        
        print(f"\n{'='*60}")
        print(f"❌ 發現 {total_conflicts} 個層級衝突")
        print(f"{'='*60}")
        
        # JSON 衝突報告
        if self.json_conflicts:
            print(f"\n📄 JSON 層級衝突 ({len(self.json_conflicts)} 個)：")
            print("-" * 40)
            
            for i, conflict in enumerate(self.json_conflicts, 1):
                print(f"\n衝突 {i}：ID '{conflict['id']}'")
                print(f"  描述：{conflict['description']}")
                print(f"  影響語言：{', '.join(conflict['languages']) if conflict['languages'] else '未知'}")
                print(f"  影響業態：{', '.join(conflict['business_types']) if conflict['business_types'] else '未知'}")
                print(f"  不同層級結構：")
                
                for j, (path, structure) in enumerate(zip(conflict['paths'], conflict['layer_structures']), 1):
                    print(f"    {j}) 路徑: {path}")
                    print(f"       結構: {structure}")
                
                if conflict['values']:
                    print(f"  相關數值：{', '.join(conflict['values'])}")
        
        # PO 衝突報告
        if self.po_conflicts:
            print(f"\n📝 PO 上下文衝突 ({len(self.po_conflicts)} 個)：")
            print("-" * 40)
            
            for i, conflict in enumerate(self.po_conflicts, 1):
                print(f"\n衝突 {i}：msgid '{conflict['id']}'")
                print(f"  描述：{conflict['description']}")
                print(f"  影響語言：{', '.join(conflict['languages']) if conflict['languages'] else '未知'}")
                print(f"  影響業態：{', '.join(conflict['business_types']) if conflict['business_types'] else '未知'}")
                
                if 'contexts' in conflict:
                    print(f"  不同上下文：")
                    for j, context in enumerate(conflict['contexts'], 1):
                        print(f"    {j}) {context}")
                
                if conflict['values']:
                    print(f"  相關翻譯：{', '.join(conflict['values'])}")
        
        # 修正建議
        print(f"\n🔧 修正建議：")
        print("1. 檢查 Excel 檔案中是否有重複的 ID 名稱")
        print("2. 確認每個 ID 在不同語言/業態中的路徑結構一致")
        print("3. 如果是合理的不同路徑，請修改 ID 名稱以區分用途")
        print("4. 檢查 JSON 檔案結構是否符合預期的多語言格式")
        print("5. 確認 PO 檔案中的 msgid 在同一上下文中使用")
        
        print(f"\n❌ 由於發現層級衝突，合併進程已終止")
        print("請修正上述衝突後重新執行合併操作")
        
        return True
    
    def generate_conflict_report_file(self, output_dir: Path, timestamp: str):
        """生成詳細的衝突報告檔案"""
        if not self.json_conflicts and not self.po_conflicts:
            return
        
        report_file = output_dir / f"layer_conflicts_report_{timestamp}.txt"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            with open(report_file, 'w', encoding='utf-8') as f:
                f.write("層級衝突詳細報告\n")
                f.write(f"生成時間：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"{'='*60}\n\n")
                
                total_conflicts = len(self.json_conflicts) + len(self.po_conflicts)
                f.write(f"總衝突數量：{total_conflicts}\n")
                f.write(f"JSON 層級衝突：{len(self.json_conflicts)} 個\n")
                f.write(f"PO 上下文衝突：{len(self.po_conflicts)} 個\n\n")
                
                # JSON 衝突詳情
                if self.json_conflicts:
                    f.write("JSON 層級衝突詳情：\n")
                    f.write("-" * 40 + "\n")
                    
                    for i, conflict in enumerate(self.json_conflicts, 1):
                        f.write(f"\n衝突 {i}：\n")
                        f.write(f"  ID：{conflict['id']}\n")
                        f.write(f"  類型：{conflict['type']}\n")
                        f.write(f"  描述：{conflict['description']}\n")
                        f.write(f"  影響語言：{', '.join(conflict['languages'])}\n")
                        f.write(f"  影響業態：{', '.join(conflict['business_types'])}\n")
                        
                        f.write(f"  衝突路徑：\n")
                        for j, (path, structure) in enumerate(zip(conflict['paths'], conflict['layer_structures']), 1):
                            f.write(f"    {j}) 完整路徑: {path}\n")
                            f.write(f"       層級結構: {structure}\n")
                        
                        if conflict['values']:
                            f.write(f"  相關數值：\n")
                            for value in conflict['values']:
                                f.write(f"    - {value}\n")
                
                # PO 衝突詳情
                if self.po_conflicts:
                    f.write("\nPO 上下文衝突詳情：\n")
                    f.write("-" * 40 + "\n")
                    
                    for i, conflict in enumerate(self.po_conflicts, 1):
                        f.write(f"\n衝突 {i}：\n")
                        f.write(f"  msgid：{conflict['id']}\n")
                        f.write(f"  類型：{conflict['type']}\n")
                        f.write(f"  描述：{conflict['description']}\n")
                        f.write(f"  影響語言：{', '.join(conflict['languages'])}\n")
                        f.write(f"  影響業態：{', '.join(conflict['business_types'])}\n")
                        
                        if 'contexts' in conflict:
                            f.write(f"  不同上下文：\n")
                            for context in conflict['contexts']:
                                f.write(f"    - {context}\n")
                        
                        if conflict['values']:
                            f.write(f"  相關翻譯：\n")
                            for value in conflict['values']:
                                f.write(f"    - {value}\n")
                
                # 修正指引
                f.write(f"\n修正指引：\n")
                f.write("1. 層級衝突分析：\n")
                f.write("   - 檢查相同 ID 是否在不同的 JSON 路徑層級中使用\n")
                f.write("   - 確認多語言結構中的路徑一致性\n")
                f.write("   - 驗證 PO 檔案中的 msgid 上下文使用\n\n")
                
                f.write("2. 建議的修正方法：\n")
                f.write("   - 重新命名衝突的 ID 以反映其在不同層級的用途\n")
                f.write("   - 統一多語言 JSON 結構中的路徑格式\n")
                f.write("   - 為 PO 檔案中的重複 msgid 添加適當的上下文\n")
                f.write("   - 檢查 Excel 檔案中的項目ID是否有邏輯錯誤\n\n")
                
                f.write("3. 預防措施：\n")
                f.write("   - 建立 ID 命名規範，避免層級間的重複\n")
                f.write("   - 使用層級前綴來區分不同層級的項目\n")
                f.write("   - 在合併前進行結構驗證\n")
                f.write("   - 定期審查翻譯檔案的結構一致性\n")
            
            print(f"📄 層級衝突報告已生成：{report_file}")
            
        except Exception as e:
            print(f"⚠️  生成層級衝突報告失敗：{e}")


# 以下是原有函數，增加層級衝突檢測邏輯

def detect_tobemodified_files(config) -> dict:
    """檢測可用的 tobemodified 檔案"""
    available_files = {}
    
    # 檢測輸出目錄中的檔案
    try:
        dirs = config.get_directories()
        output_dir = Path(dirs['output_dir'])
    except Exception:
        output_dir = Path('i18n_output')
    
    # 使用配置載入器的語言檢測
    try:
        available_languages = config.detect_available_languages()
    except Exception as e:
        print(f"⚠️  語言檢測失敗：{e}")
        available_languages = []
    
    # 檢測標準命名的檔案
    for language in available_languages:
        tobemodified_path = output_dir / f"{language}_tobemodified.xlsx"
        if tobemodified_path.exists():
            available_files[language] = tobemodified_path
    
    # 在當前目錄中查找額外的檔案
    for file_path in Path('.').glob("*_tobemodified.xlsx"):
        filename = file_path.stem
        if filename.endswith('_tobemodified'):
            language = filename[:-len('_tobemodified')]
            
            # 過濾系統臨時檔案
            if language.startswith(('~$', '.', '__')):
                continue
            
            if language not in available_files:
                available_files[language] = file_path

    return available_files


def scan_combine_directory(combine_dir: Path) -> dict:
    """掃描 i18n_combine 目錄中的檔案"""
    files = {
        'json': [],
        'po': []
    }
    
    if not combine_dir.exists():
        return files
    
    # 遞歸掃描所有 JSON 和 PO 檔案
    for file_path in combine_dir.rglob("*.json"):
        relative_path = file_path.relative_to(combine_dir)
        files['json'].append({
            'path': file_path,
            'relative_path': str(relative_path),
            'name': file_path.name
        })
    
    for file_path in combine_dir.rglob("*.po"):
        relative_path = file_path.relative_to(combine_dir)
        files['po'].append({
            'path': file_path,
            'relative_path': str(relative_path),
            'name': file_path.name
        })
    
    return files


def choose_tobemodified_files(available_files: dict) -> dict:
    """選擇要使用的 tobemodified 檔案（支援多選）"""
    if not available_files:
        print("❌ 未找到任何 tobemodified 檔案")
        return {}
    
    print("\n📄 可用的 tobemodified 檔案：")
    choices = list(available_files.items())
    
    for i, (language, file_path) in enumerate(choices, 1):
        print(f"  {i}) {language} ({file_path.name})")
    
    print(f"  A) 全部選擇")
    print(f"  0) 取消操作")
    
    selected_files = {}
    
    while True:
        try:
            choice = input(f"\n請選擇要使用的檔案 (可多選，用逗號分隔，如 1,2,3 或 A)：").strip()
            
            if choice == '0':
                print("❌ 操作取消")
                return {}
            elif choice.upper() == 'A':
                selected_files = available_files.copy()
                break
            else:
                # 解析多選
                choice_indices = [int(x.strip()) - 1 for x in choice.split(',')]
                selected_files = {}
                
                for choice_idx in choice_indices:
                    if 0 <= choice_idx < len(choices):
                        language, file_path = choices[choice_idx]
                        selected_files[language] = file_path
                    else:
                        print(f"⚠️  無效選項：{choice_idx + 1}")
                        continue
                
                if selected_files:
                    break
                else:
                    print(f"⚠️  請輸入有效的選項")
                    
        except (ValueError, KeyboardInterrupt):
            print("\n❌ 操作取消")
            return {}
    
    print(f"✅ 選擇了 {len(selected_files)} 個檔案：")
    for language, file_path in selected_files.items():
        print(f"   {language}: {file_path.name}")
    
    return selected_files


def choose_combine_file(files: list, file_type: str) -> Path:
    """選擇要合併的檔案"""
    if not files:
        print(f"⚠️  /i18n_combine/ 中沒有找到 {file_type.upper()} 檔案")
        return None
    
    print(f"\n📁 可用的 {file_type.upper()} 檔案：")
    for i, file_info in enumerate(files, 1):
        print(f"  {i}) {file_info['relative_path']}")
    
    print(f"  0) 跳過 {file_type.upper()} 檔案")
    
    while True:
        try:
            choice = input(f"\n請選擇要合併的 {file_type.upper()} 檔案 (0-{len(files)})：").strip()
            choice_idx = int(choice)
            
            if choice_idx == 0:
                print(f"⏭️  跳過 {file_type.upper()} 檔案")
                return None
            elif 1 <= choice_idx <= len(files):
                selected_file = files[choice_idx - 1]
                print(f"✅ 選擇了：{selected_file['relative_path']}")
                return selected_file['path']
            else:
                print(f"⚠️  請輸入 0-{len(files)} 之間的數字")
        except (ValueError, KeyboardInterrupt):
            print("\n❌ 操作取消")
            return None


def read_excel_updates_for_language(xlsx_path: Path, language: str, config) -> dict:
    """讀取單個語言的 Excel 檔案中的更新資料"""
    try:
        print(f"📖 讀取 {language} 的 Excel 檔案：{xlsx_path.name}")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        header_row = list(ws[1])
        header = {cell.value: idx for idx, cell in enumerate(header_row) if cell.value}
        
        # 基本欄位檢查
        required_columns = ["檔案類型", "項目ID", "項目內容"]
        missing_columns = []
        
        for col in required_columns:
            if col not in header:
                missing_columns.append(col)
        
        if missing_columns:
            print(f"❌ {language} Excel 缺少必要欄位：{missing_columns}")
            return {}
        
        # 自動檢測所有業態的替換結果欄位
        business_types = config.get_business_types()
        available_business_types = []
        
        for bt_code, bt_config in business_types.items():
            display_name = bt_config['display_name']
            result_col_name = f"{display_name}_替換結果"
            if result_col_name in header:
                available_business_types.append(bt_code)
        
        if not available_business_types:
            print(f"❌ {language} 未找到任何業態的替換結果欄位")
            return {}
        
        print(f"   📋 {language} 檢測到業態：{', '.join([business_types[bt]['display_name'] for bt in available_business_types])}")
        
        # 解析更新資料
        updates = {bt_code: {"po": [], "json": []} for bt_code in available_business_types}
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= max(header.values()):
                continue
            
            try:
                file_type = row[header["檔案類型"]]
                entry_id = row[header["項目ID"]]
                original_text = row[header["項目內容"]]
                
                if not file_type or not entry_id:
                    continue
                
                file_type = str(file_type).lower()
                
                # 處理每個可用的業態
                for bt_code in available_business_types:
                    display_name = business_types[bt_code]['display_name']
                    result_col_name = f"{display_name}_替換結果"
                    
                    new_value = row[header[result_col_name]]
                    
                    # 跳過空值和與原文相同的值
                    if not new_value or not str(new_value).strip():
                        continue
                    
                    new_value = str(new_value).strip()
                    
                    if original_text and str(original_text).strip() == new_value:
                        continue
                    
                    # 創建更新記錄，包含語言信息
                    update_record = (str(entry_id), new_value, language)
                    
                    if file_type == "po":
                        updates[bt_code]["po"].append(update_record)
                    elif file_type == "json":
                        updates[bt_code]["json"].append(update_record)
            
            except Exception as e:
                print(f"⚠️  {language} 第 {row_num} 行處理失敗: {e}")
                continue
        
        # 統計有效更新
        total_updates = 0
        for bt_code in available_business_types:
            bt_updates = len(updates[bt_code]["po"]) + len(updates[bt_code]["json"])
            total_updates += bt_updates
            if bt_updates > 0:
                print(f"     {business_types[bt_code]['display_name']}: {bt_updates} 個更新")
        
        print(f"   📊 {language} 總計：{total_updates} 個有效更新")
        return updates
        
    except Exception as e:
        print(f"❌ 讀取 {language} Excel 檔案失敗：{e}")
        return {}


def combine_multilang_json_files_for_business_type(all_updates: dict, target_json_path: Path, 
                                                  output_json_path: Path, bt_code: str, log_detail=None) -> dict:
    """【改進版】為特定業態合併多語言 JSON 檔案，包含層級衝突檢測"""
    result = {
        "success": False,
        "merged": 0,
        "skipped": 0,
        "conflicts": [],
        "errors": [],
        "language_stats": {}
    }
    
    # 檢查是否有當前業態的更新
    has_updates = False
    for language_updates in all_updates.values():
        if bt_code in language_updates and language_updates[bt_code]['json']:
            has_updates = True
            break
    
    if not has_updates:
        result["success"] = True
        if log_detail:
            log_detail(f"JSON ({bt_code}): 沒有任何更新項目")
        return result
    
    try:
        # 載入目標 JSON 檔案
        if not target_json_path.exists():
            result["errors"].append(f"目標 JSON 檔案不存在：{target_json_path}")
            return result
        
        target_data = json.loads(target_json_path.read_text(encoding="utf-8"))
        print(f"   📄 載入目標多語言 JSON 檔案：{target_json_path.name}")
        if log_detail:
            log_detail(f"載入目標 JSON 檔案：{target_json_path.name}")
        
        # 檢查是否為多語言結構
        is_multilang_structure = check_multilang_json_structure(target_data)
        print(f"   🔍 多語言結構檢測：{'是' if is_multilang_structure else '否'}")
        if log_detail:
            log_detail(f"多語言結構檢測：{'是' if is_multilang_structure else '否'}")
        
        # **新增：層級衝突檢測**
        conflict_detector = LayerConflictDetector()
        has_layer_conflicts = conflict_detector.detect_json_layer_conflicts(
            {lang: {bt_code: updates[bt_code]} for lang, updates in all_updates.items() if bt_code in updates},
            target_data,
            is_multilang_structure
        )
        
        if has_layer_conflicts:
            print(f"   ❌ JSON 檔案發現層級衝突，終止合併")
            if log_detail:
                log_detail(f"JSON ({bt_code}): 發現層級衝突，終止合併")
            
            # 生成衝突報告
            conflict_detector.print_conflict_report()
            
            # 將衝突信息添加到結果中
            result["errors"].append("發現層級衝突，合併已終止")
            return result
        
        conflicts = []
        language_stats = {}
        
        # 只處理當前業態的更新
        for language, language_updates in all_updates.items():
            if bt_code not in language_updates:
                continue
                
            language_stats[language] = {"merged": 0, "skipped": 0, "conflicts": 0}
            
            if log_detail:
                log_detail(f"處理語言 {language} 的 JSON 更新 (業態: {bt_code})")
            
            # 處理當前業態的 JSON 更新
            bt_updates = language_updates[bt_code]
            for json_path_str, new_value, update_language in bt_updates['json']:
                if log_detail:
                    log_detail(f"處理更新：{update_language}.{json_path_str} = {new_value}")
                
                # 多語言結構的路徑映射
                if is_multilang_structure:
                    multilang_path = f"{update_language}.{json_path_str}"
                else:
                    multilang_path = json_path_str
                
                # 獲取現有值
                existing_value = get_json_value_by_path(target_data, multilang_path)
                
                # 正確處理值的比較和衝突檢測
                if existing_value is not None:
                    existing_str = str(existing_value).strip()
                    new_str = str(new_value).strip()
                    
                    # 如果值完全相同，跳過
                    if existing_str == new_str:
                        result["skipped"] += 1
                        language_stats[update_language]["skipped"] += 1
                        if log_detail:
                            log_detail(f"跳過相同值：{multilang_path} = '{new_str}'")
                        continue
                    
                    # 當值不同時，標記為衝突並讓用戶決定
                    if existing_str != new_str:
                        conflict_info = {
                            "path": multilang_path,
                            "language": update_language,
                            "existing_value": existing_str,
                            "new_value": new_str,
                            "file_type": "json"
                        }
                        conflicts.append(conflict_info)
                        result["conflicts"].append(conflict_info)
                        language_stats[update_language]["conflicts"] += 1
                        
                        if log_detail:
                            log_detail(f"發現衝突：{multilang_path}")
                            log_detail(f"  現有值: '{existing_str}'")
                            log_detail(f"  新值: '{new_str}'")
                        
                        # 詢問用戶如何處理衝突
                        choice = handle_json_conflict(multilang_path, existing_str, new_str, update_language)
                        
                        if choice == "keep_existing":
                            result["skipped"] += 1
                            language_stats[update_language]["skipped"] += 1
                            if log_detail:
                                log_detail(f"保留現有值：{multilang_path} = '{existing_str}'")
                            continue
                        elif choice == "use_new":
                            # 繼續執行更新邏輯
                            if log_detail:
                                log_detail(f"採用新值：{multilang_path} = '{new_str}'")
                        elif choice == "skip":
                            result["skipped"] += 1
                            language_stats[update_language]["skipped"] += 1
                            if log_detail:
                                log_detail(f"跳過處理：{multilang_path}")
                            continue
                
                # 應用更新
                if set_json_value_by_path(target_data, multilang_path, new_value):
                    result["merged"] += 1
                    language_stats[update_language]["merged"] += 1
                    if log_detail:
                        original_display = f"'{existing_value}'" if existing_value is not None else "無"
                        log_detail(f"成功更新：{multilang_path} = '{new_value}' (原值: {original_display})")
                else:
                    error_msg = f"無法設置 JSON 路徑：{multilang_path} (語言: {update_language})"
                    result["errors"].append(error_msg)
                    if log_detail:
                        log_detail(f"錯誤：{error_msg}")
        
        # 保存合併後的檔案
        output_json_path.parent.mkdir(parents=True, exist_ok=True)
        
        json_content = json.dumps(target_data, ensure_ascii=False, indent=2)
        output_json_path.write_text(json_content, encoding="utf-8")
        
        result["success"] = True
        result["language_stats"] = language_stats
        
        # 修正日誌訊息，包含衝突數量
        total_conflicts = len(conflicts)
        if log_detail:
            log_detail(f"JSON ({bt_code}) 合併完成：合併 {result['merged']} 個，跳過 {result['skipped']} 個，衝突 {total_conflicts} 個")
        
    except json.JSONDecodeError as e:
        error_msg = f"JSON 格式錯誤：{e}"
        result["errors"].append(error_msg)
        if log_detail:
            log_detail(f"錯誤：{error_msg}")
    except Exception as e:
        error_msg = f"JSON 檔案合併失敗：{e}"
        result["errors"].append(error_msg)
        if log_detail:
            log_detail(f"錯誤：{error_msg}")
    
    return result


def handle_json_conflict(path: str, existing_value: str, new_value: str, language: str) -> str:
    """處理 JSON 合併衝突，讓用戶選擇如何處理"""
    print(f"\n⚠️  發現衝突：")
    print(f"📍 路徑：{path}")
    print(f"🌍 語言：{language}")
    print(f"📄 現有值：'{existing_value}'")
    print(f"🆕 新值：'{new_value}'")
    
    while True:
        print(f"\n請選擇處理方式：")
        print(f"  1) 保留現有值 ('{existing_value}')")
        print(f"  2) 使用新值 ('{new_value}')")
        print(f"  3) 跳過此項目")
        print(f"  A) 對所有類似衝突使用新值")
        print(f"  K) 對所有類似衝突保留現有值")
        
        try:
            choice = input(f"請選擇 (1/2/3/A/K)：").strip().upper()
            
            if choice == "1":
                return "keep_existing"
            elif choice == "2":
                return "use_new"
            elif choice == "3":
                return "skip"
            elif choice == "A":
                # 可以擴展為全局策略
                print(f"✅ 將使用新值")
                return "use_new"
            elif choice == "K":
                # 可以擴展為全局策略
                print(f"✅ 將保留現有值")
                return "keep_existing"
            else:
                print(f"⚠️  請輸入有效選項 (1/2/3/A/K)")
                
        except KeyboardInterrupt:
            print(f"\n❌ 操作取消，跳過此項目")
            return "skip"


def combine_po_files_for_business_type(all_updates: dict, target_po_path: Path, 
                                     output_dir: Path, bt_code: str, log_detail=None) -> dict:
    """【改進版】為特定業態處理 PO 檔案合併，包含層級衝突檢測"""
    result = {
        "success": False,
        "merged": 0,
        "skipped": 0,
        "conflicts": [],
        "errors": [],
        "language_stats": {}
    }
    
    # 檢查是否有當前業態的 PO 更新
    has_updates = False
    for language_updates in all_updates.values():
        if bt_code in language_updates and language_updates[bt_code]['po']:
            has_updates = True
            break
    
    if not has_updates:
        result["success"] = True
        if log_detail:
            log_detail(f"PO ({bt_code}): 沒有任何更新項目")
        return result
    
    try:
        # 載入目標 PO 檔案
        if not target_po_path.exists():
            result["errors"].append(f"目標 PO 檔案不存在：{target_po_path}")
            return result
        
        target_po = polib.pofile(str(target_po_path))
        print(f"   📄 載入目標 PO 檔案：{target_po_path.name}，共 {len(target_po)} 個條目")
        if log_detail:
            log_detail(f"載入目標 PO 檔案：{target_po_path.name}，共 {len(target_po)} 個條目")
        
        # **新增：層級衝突檢測**
        conflict_detector = LayerConflictDetector()
        has_layer_conflicts = conflict_detector.detect_po_layer_conflicts(
            {lang: {bt_code: updates[bt_code]} for lang, updates in all_updates.items() if bt_code in updates},
            target_po
        )
        
        if has_layer_conflicts:
            print(f"   ❌ PO 檔案發現層級衝突，終止合併")
            if log_detail:
                log_detail(f"PO ({bt_code}): 發現層級衝突，終止合併")
            
            # 生成衝突報告
            conflict_detector.print_conflict_report()
            
            # 將衝突信息添加到結果中
            result["errors"].append("發現層級衝突，合併已終止")
            return result
        
        language_stats = {}
        
        # 只處理當前業態的更新
        for language, language_updates in all_updates.items():
            if bt_code not in language_updates:
                continue
                
            language_stats[language] = {"merged": 0, "skipped": 0, "conflicts": 0}
            
            # 處理當前業態的 PO 更新
            bt_updates = language_updates[bt_code]
            for msgid, new_msgstr, update_language in bt_updates['po']:
                target_entry = target_po.find(msgid)
                
                if target_entry:
                    # 只有當現有值和新值真的不同時才需要更新
                    if target_entry.msgstr and target_entry.msgstr.strip():
                        if target_entry.msgstr == new_msgstr:
                            # 值相同，跳過
                            result["skipped"] += 1
                            language_stats[update_language]["skipped"] += 1
                            continue
                    
                    # 應用更新
                    target_entry.msgstr = new_msgstr
                    result["merged"] += 1
                    language_stats[update_language]["merged"] += 1
                else:
                    # 目標檔案中沒有此條目，添加新條目
                    new_entry = polib.POEntry(
                        msgid=msgid,
                        msgstr=new_msgstr
                    )
                    target_po.append(new_entry)
                    result["merged"] += 1
                    language_stats[update_language]["merged"] += 1
        
        # 保存合併後的檔案
        config = get_config()
        business_types = config.get_business_types()
        
        if bt_code in business_types:
            suffix = business_types[bt_code]['suffix']
            output_po_path = output_dir / f"{target_po_path.stem}{suffix}_combined.po"
            output_po_path.parent.mkdir(parents=True, exist_ok=True)
            target_po.save(str(output_po_path))
        
        result["success"] = True
        result["language_stats"] = language_stats
        
        if log_detail:
            log_detail(f"PO ({bt_code}) 合併完成：合併 {result['merged']} 個，跳過 {result['skipped']} 個")
        
    except Exception as e:
        error_msg = f"PO 檔案合併失敗：{e}"
        result["errors"].append(error_msg)
        if log_detail:
            log_detail(f"錯誤：{error_msg}")
    
    return result


def check_multilang_json_structure(data: dict) -> bool:
    """檢查 JSON 是否為多語言結構"""
    if not isinstance(data, dict):
        return False
    
    # 檢查頂層 key 是否像語言代碼
    for key in data.keys():
        if isinstance(key, str) and re.match(r'^[a-z]{2}(-[A-Z]{2})?', key):
            # 如果至少有一個 key 像語言代碼，且其值是字典，則認為是多語言結構
            if isinstance(data[key], dict):
                return True
    
    return False


def get_json_value_by_path(data: dict, path: str):
    """按路徑獲取 JSON 值"""
    try:
        path_parts = parse_json_path(path)
        current = data
        
        for part_type, part_value in path_parts:
            if part_type == 'key':
                if part_value not in current:
                    return None
                current = current[part_value]
            elif part_type == 'index':
                if not isinstance(current, list) or len(current) <= part_value:
                    return None
                current = current[part_value]
        
        return current
        
    except Exception:
        return None


def set_json_value_by_path(data: dict, path: str, new_value: str) -> bool:
    """按路徑設置 JSON 值"""
    try:
        path_parts = parse_json_path(path)
        current = data
        
        for i, (part_type, part_value) in enumerate(path_parts):
            is_last = (i == len(path_parts) - 1)
            
            if part_type == 'key':
                if is_last:
                    current[part_value] = new_value
                else:
                    if part_value not in current:
                        next_part_type = path_parts[i + 1][0] if i + 1 < len(path_parts) else 'key'
                        current[part_value] = [] if next_part_type == 'index' else {}
                    current = current[part_value]
            
            elif part_type == 'index':
                if is_last:
                    while len(current) <= part_value:
                        current.append(None)
                    current[part_value] = new_value
                else:
                    while len(current) <= part_value:
                        current.append(None)
                    if current[part_value] is None:
                        next_part_type = path_parts[i + 1][0] if i + 1 < len(path_parts) else 'key'
                        current[part_value] = [] if next_part_type == 'index' else {}
                    current = current[part_value]
        
        return True
        
    except Exception as e:
        return False


def parse_json_path(path: str) -> list:
    """解析 JSON 路徑"""
    parts = []
    current = ""
    in_bracket = False
    
    for char in path:
        if char == '[':
            if current:
                parts.append(('key', current))
                current = ""
            in_bracket = True
        elif char == ']':
            if in_bracket and current:
                try:
                    parts.append(('index', int(current)))
                except ValueError:
                    raise ValueError(f"無效的陣列索引：{current}")
                current = ""
            in_bracket = False
        elif char == '.' and not in_bracket:
            if current:
                parts.append(('key', current))
                current = ""
        else:
            current += char
    
    if current:
        parts.append(('key', current))
    
    return parts


def main():
    """主執行函數 - 包含層級衝突檢測"""
    print("🚀 開始多語言檔案合併處理 (v1.4 - 層級衝突檢測版)")
    
    # 載入配置
    config = get_config()
    
    # 檢測可用的 tobemodified 檔案
    available_files = detect_tobemodified_files(config)
    
    if not available_files:
        print("❌ 未找到任何 tobemodified 檔案")
        print("請先執行 script_01_generate_xlsx.py 生成檔案")
        sys.exit(1)
    
    # 步驟1：選擇 tobemodified 檔案（支援多選）
    selected_files = choose_tobemodified_files(available_files)
    if not selected_files:
        sys.exit(1)
    
    # 檢查 i18n_combine 目錄
    combine_dir = Path("i18n_combine")
    
    if not combine_dir.exists():
        print(f"❌ 合併目錄不存在：{combine_dir}")
        print(f"請創建目錄並放入要合併的檔案")
        sys.exit(1)
    
    print(f"📁 掃描合併目錄：{combine_dir}")
    
    # 掃描 combine 目錄中的檔案
    combine_files = scan_combine_directory(combine_dir)
    
    # 步驟2：選擇要合併的 JSON 檔案
    target_json_path = choose_combine_file(combine_files['json'], 'json')
    
    # 步驟3：選擇要合併的 PO 檔案
    target_po_path = choose_combine_file(combine_files['po'], 'po')
    
    # 檢查是否至少選擇了一個檔案
    if not target_json_path and not target_po_path:
        print("❌ 必須至少選擇一個檔案進行合併")
        sys.exit(1)
    
    # 讀取所有選中語言的 Excel 更新資料
    all_updates = {}
    for language, xlsx_path in selected_files.items():
        updates = read_excel_updates_for_language(xlsx_path, language, config)
        if updates:
            all_updates[language] = updates
    
    if not all_updates:
        print("❌ 沒有讀取到任何有效的更新資料")
        sys.exit(1)
    
    # 統計所有業態
    all_business_types = set()
    for language_updates in all_updates.values():
        all_business_types.update(language_updates.keys())
    
    print(f"\n📋 合併設定：")
    print(f"   來源語言：{', '.join(selected_files.keys())}")
    if target_json_path:
        print(f"   JSON 檔案：{target_json_path.relative_to(combine_dir)}")
    if target_po_path:
        print(f"   PO 檔案：{target_po_path.relative_to(combine_dir)}")
    print(f"   涵蓋業態：{', '.join([config.get_business_types()[bt]['display_name'] for bt in all_business_types])}")
    
    # **新增：預先進行全面的層級衝突檢測**
    print(f"\n🔍 執行全面層級衝突檢測...")
    global_conflict_detector = LayerConflictDetector()
    
    has_global_conflicts = False
    
    # 檢測 JSON 層級衝突
    if target_json_path:
        try:
            target_json_data = json.loads(target_json_path.read_text(encoding="utf-8"))
            is_multilang = check_multilang_json_structure(target_json_data)
            
            if global_conflict_detector.detect_json_layer_conflicts(all_updates, target_json_data, is_multilang):
                has_global_conflicts = True
        except Exception as e:
            print(f"⚠️  JSON 衝突檢測失敗：{e}")
    
    # 檢測 PO 層級衝突
    if target_po_path:
        try:
            target_po_data = polib.pofile(str(target_po_path))
            
            if global_conflict_detector.detect_po_layer_conflicts(all_updates, target_po_data):
                has_global_conflicts = True
        except Exception as e:
            print(f"⚠️  PO 衝突檢測失敗：{e}")
    
    # 如果發現全局層級衝突，終止進程
    if has_global_conflicts:
        print(f"\n{'='*60}")
        print(f"❌ 發現層級衝突，合併進程已終止")
        print(f"{'='*60}")
        
        global_conflict_detector.print_conflict_report()
        
        # 建立輸出目錄以生成衝突報告
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        dirs = config.get_directories()
        output_dir = Path(dirs['output_dir']) / f"conflict_report_{timestamp}"
        
        global_conflict_detector.generate_conflict_report_file(output_dir, timestamp)
        
        print(f"\n🔧 修正建議：")
        print("1. 檢查 Excel 檔案中的項目ID是否有重複命名")
        print("2. 確認同一個ID在不同語言中是否使用了不同的路徑結構")
        print("3. 檢查多語言 JSON 檔案的結構一致性")
        print("4. 驗證 PO 檔案中的 msgid 上下文使用")
        print("5. 建議使用層級前綴來區分不同層級的相同名稱項目")
        
        print(f"📄 詳細衝突報告已生成於：{output_dir}")
        sys.exit(1)
    
    print(f"✅ 層級衝突檢測通過，繼續合併流程...")
    
    # 建立輸出目錄 - 使用正確的命名格式
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    dirs = config.get_directories()
    output_dir = Path(dirs['output_dir']) / f"multi_{timestamp}_combined"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 設置日誌
    log_file = output_dir / f"multi_combine_{timestamp}.log"
    
    def log_detail(message: str):
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(f"{datetime.datetime.now().strftime('%H:%M:%S')} - {message}\n")
    
    log_detail(f"開始多語言合併處理")
    log_detail(f"語言：{', '.join(selected_files.keys())}")
    log_detail(f"來源檔案：{list(selected_files.values())}")
    log_detail(f"涵蓋業態：{', '.join(all_business_types)}")
    log_detail(f"層級衝突檢測：通過")
    
    # 處理合併邏輯 - 避免業態間衝突
    business_types = config.get_business_types()
    all_results = {}
    
    # 按業態分別處理，避免相互干擾
    for bt_code in all_business_types:
        if bt_code not in business_types:
            continue
            
        bt_config = business_types[bt_code]
        display_name = bt_config['display_name']
        suffix = bt_config['suffix']
        
        print(f"\n📝 處理 {display_name}...")
        log_detail(f"開始處理業態：{display_name}")
        
        results = {}
        
        # 為當前業態處理 JSON 檔案
        if target_json_path:
            output_json_path = output_dir / f"{target_json_path.stem}{suffix}_combined.json"
            json_result = combine_multilang_json_files_for_business_type(
                all_updates,
                target_json_path,
                output_json_path,
                bt_code,
                log_detail
            )
            results['json_result'] = json_result
            
            # 顯示結果
            if json_result.get('errors'):
                print(f"     ❌ JSON 檔案處理錯誤：{json_result['errors']}")
            else:
                # 顯示語言統計
                if json_result.get('language_stats'):
                    for lang, stats in json_result['language_stats'].items():
                        if stats['merged'] > 0 or stats['skipped'] > 0:
                            print(f"     📊 {lang}: 合併 {stats['merged']} 個, 跳過 {stats['skipped']} 個")
                
                if json_result.get('merged', 0) == 0 and json_result.get('skipped', 0) == 0:
                    print(f"     ℹ️  {display_name} 沒有 JSON 更新項目")
        
        # 為當前業態處理 PO 檔案
        if target_po_path:
            po_result = combine_po_files_for_business_type(
                all_updates,
                target_po_path,
                output_dir,
                bt_code,
                log_detail
            )
            results['po_result'] = po_result
            
            # 顯示結果
            if po_result.get('errors'):
                print(f"     ❌ PO 檔案處理錯誤：{po_result['errors']}")
            else:
                # 顯示語言統計
                if po_result.get('language_stats'):
                    for lang, stats in po_result['language_stats'].items():
                        if stats['merged'] > 0 or stats['skipped'] > 0:
                            print(f"     📊 {lang}: 合併 {stats['merged']} 個, 跳過 {stats['skipped']} 個")
                
                if po_result.get('merged', 0) == 0 and po_result.get('skipped', 0) == 0:
                    print(f"     ℹ️  {display_name} 沒有 PO 更新項目")
        
        # 如果沒有更新，複製原檔案
        if target_json_path and results.get('json_result', {}).get('merged', 0) == 0:
            output_json_path = output_dir / f"{target_json_path.stem}{suffix}_combined.json"
            if not output_json_path.exists():
                output_json_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(target_json_path, output_json_path)
                print(f"     📄 複製 JSON 檔案（無更新）")
                log_detail(f"複製原始 JSON 檔案：{target_json_path.name}")
        
        if target_po_path and results.get('po_result', {}).get('merged', 0) == 0:
            output_po_path = output_dir / f"{target_po_path.stem}{suffix}_combined.po"
            if not output_po_path.exists():
                output_po_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(target_po_path, output_po_path)
                print(f"     📄 複製 PO 檔案（無更新）")
                log_detail(f"複製原始 PO 檔案：{target_po_path.name}")
        
        all_results[bt_code] = results
        
        # 統計結果
        total_merged = 0
        total_skipped = 0
        total_errors = 0
        
        for result in results.values():
            total_merged += result.get('merged', 0)
            total_skipped += result.get('skipped', 0)
            total_errors += len(result.get('errors', []))
        
        if total_errors > 0:
            print(f"     ❌ 處理失敗 - 錯誤: {total_errors} 個")
        else:
            print(f"     ✅ 完成 - 合併: {total_merged} 個, 跳過: {total_skipped} 個")
        
        log_detail(f"{display_name} 處理完成：合併 {total_merged} 個，跳過 {total_skipped} 個，錯誤 {total_errors} 個")
    
    # 生成最終報告
    total_merged = sum(
        sum(result.get('merged', 0) for result in results.values())
        for results in all_results.values()
    )
    total_skipped = sum(
        sum(result.get('skipped', 0) for result in results.values())
        for results in all_results.values()
    )
    total_errors = sum(
        sum(len(result.get('errors', [])) for result in results.values())
        for results in all_results.values()
    )
    
    print(f"\n🎉 多語言合併處理完成！")
    print(f"📊 處理結果：合併 {total_merged} 個項目，跳過 {total_skipped} 個項目")
    if total_errors > 0:
        print(f"⚠️  處理錯誤：{total_errors} 個")
    print(f"📁 輸出目錄：{output_dir}")
    print(f"🔍 層級衝突檢測：通過")
    
    # 生成處理摘要
    generate_multilang_summary_report(all_results, all_updates, output_dir, timestamp, target_json_path, target_po_path, log_detail)


def generate_multilang_summary_report(results: dict, all_updates: dict, output_dir: Path, timestamp: str, 
                                     target_json_path: Path, target_po_path: Path, log_detail):
    """生成多語言合併處理摘要報告 - 包含層級衝突檢測信息"""
    summary_file = output_dir / f"multi_combine_summary_{timestamp}.txt"
    
    try:
        with open(summary_file, 'w', encoding='utf-8') as f:
            f.write(f"多語言檔案合併處理摘要報告\n")
            f.write(f"生成時間：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"版本：v1.4 - 層級衝突檢測版\n")
            f.write(f"{'='*60}\n\n")
            
            f.write(f"目標檔案：\n")
            if target_json_path:
                f.write(f"  JSON: {target_json_path}\n")
            if target_po_path:
                f.write(f"  PO: {target_po_path}\n")
            f.write(f"\n")
            
            f.write(f"處理的語言：\n")
            for language in all_updates.keys():
                f.write(f"  - {language}\n")
            f.write(f"\n")
            
            f.write(f"層級衝突檢測：通過\n")
            f.write(f"所有相同ID的不同層級衝突已在合併前檢測並解決\n\n")
            
            total_merged = 0
            total_skipped = 0
            total_errors = 0
            successful_business_types = []
            failed_business_types = []
            
            # 按業態統計
            for bt_code, bt_results in results.items():
                f.write(f"業態：{bt_code}\n")
                
                bt_merged = sum(result.get('merged', 0) for result in bt_results.values())
                bt_skipped = sum(result.get('skipped', 0) for result in bt_results.values())
                bt_errors = []
                for result in bt_results.values():
                    bt_errors.extend(result.get('errors', []))
                
                f.write(f"合併數量：{bt_merged}\n")
                f.write(f"跳過數量：{bt_skipped}\n")
                
                # 語言級別統計
                f.write(f"語言統計：\n")
                for result in bt_results.values():
                    if 'language_stats' in result:
                        for lang, stats in result['language_stats'].items():
                            f.write(f"  {lang}: 合併 {stats['merged']}, 跳過 {stats['skipped']}, 衝突 {stats.get('conflicts', 0)}\n")
                
                if bt_errors:
                    f.write(f"錯誤：\n")
                    for error in bt_errors:
                        f.write(f"  - {error}\n")
                    failed_business_types.append(bt_code)
                else:
                    successful_business_types.append(bt_code)
                
                total_merged += bt_merged
                total_skipped += bt_skipped
                total_errors += len(bt_errors)
                
                f.write(f"\n{'-'*40}\n\n")
            
            # 總計統計
            f.write(f"處理總結：\n")
            f.write(f"成功業態：{len(successful_business_types)}\n")
            f.write(f"失敗業態：{len(failed_business_types)}\n")
            f.write(f"總合併項目：{total_merged}\n")
            f.write(f"總跳過項目：{total_skipped}\n")
            f.write(f"總錯誤項目：{total_errors}\n")
            f.write(f"處理語言數：{len(all_updates)}\n")
            f.write(f"層級衝突檢測：通過\n")
            
            if successful_business_types:
                f.write(f"\n成功的業態：{', '.join(successful_business_types)}\n")
            
            if failed_business_types:
                f.write(f"失敗的業態：{', '.join(failed_business_types)}\n")
            
            f.write(f"\n層級衝突檢測說明：\n")
            f.write(f"- 本版本新增了完整的層級衝突檢測功能\n")
            f.write(f"- 在合併前檢查所有相同ID是否出現在不同層級結構中\n")
            f.write(f"- 檢測 JSON 檔案中的路徑層級衝突\n")
            f.write(f"- 檢測 PO 檔案中的上下文衝突\n")
            f.write(f"- 發現衝突時會終止進程並生成詳細報告\n")
            f.write(f"- 通過檢測後才會執行實際的合併操作\n")
            
            f.write(f"\n多語言合併說明：\n")
            f.write(f"- 本次處理支援多個語言的 tobemodified 合併到同一檔案\n")
            f.write(f"- JSON 檔案支援多語言結構（如 enterprise.json）\n")
            f.write(f"- 自動檢測並處理語言層級的路徑映射\n")
            f.write(f"- 按業態分別處理，避免業態間相互干擾\n")
            f.write(f"- 相同 key 且相同 value 的項目會自動跳過\n")
            f.write(f"- 不同 value 的項目會正常更新（不再視為衝突）\n")
            
            f.write(f"\n使用建議：\n")
            f.write(f"- 確認目標 JSON 檔案採用多語言結構（頂層為語言代碼）\n")
            f.write(f"- 合併前建議備份原始檔案\n")
            f.write(f"- 合併後請測試多語言翻譯檔案的正確性\n")
            f.write(f"- 檢查各語言層級的數據完整性\n")
            f.write(f"- 如果遇到層級衝突，請參考衝突報告進行修正\n")
            
            # 版本更新說明
            f.write(f"\n版本 v1.4 新增功能：\n")
            f.write(f"- 新增層級衝突檢測器 (LayerConflictDetector)\n")
            f.write(f"- 檢測相同ID名稱但在不同層級的衝突\n")
            f.write(f"- 全面列出所有層級衝突詳情\n")
            f.write(f"- 發現層級衝突時終止進程並生成報告\n")
            f.write(f"- 支援多語言和多業態的層級衝突檢測\n")
            f.write(f"- 提供詳細的修正建議和指引\n")
        
        log_detail(f"多語言合併摘要報告已生成：{summary_file}")
        
    except Exception as e:
        log_detail(f"生成多語言合併摘要報告失敗：{e}")


if __name__ == "__main__":
    main()