#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
script_02_apply_fixes_partial.py (v1.0 - 部分檔案生成版)

修改內容：
1. ✅ 只生成有修改的項目到部分檔案（messages_partial.po 和 {language}_partial.json）
2. ✅ 保持原有檔案結構和層級，確保 i18n 相容性
3. ✅ 支援多重敏感詞和新格式
4. ✅ 自動跳過空的替換結果，避免無意義的處理
5. ✅ 生成詳細的處理報告

依據各語言的 tobemodified_{language}.xlsx，僅將有修正的項目寫入部分檔案
"""

import json
import sys
import datetime
import argparse
import glob
from pathlib import Path
from collections import defaultdict
from config_loader import get_config

try:
    import openpyxl
    import polib
except ImportError as e:
    print(f"❌ 缺少必要套件：{e}")
    print("請執行：pip install openpyxl polib")
    sys.exit(1)


def read_and_validate_xlsx(xlsx_path: Path, config, target_business_types: list, log_detail) -> tuple:
    """讀取並驗證 Excel 檔案"""
    try:
        log_detail(f"開始讀取 Excel 檔案: {xlsx_path}")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        header_row = list(ws[1])
        header = {cell.value: idx for idx, cell in enumerate(header_row) if cell.value}
        
        log_detail(f"發現欄位: {list(header.keys())}")
        
        # 基本欄位檢查
        required_columns = ["檔案類型", "項目ID", "項目內容", "敏感詞"]
        missing_columns = []
        
        for col in required_columns:
            if col not in header:
                missing_columns.append(col)
        
        # 檢查業態替換結果欄位
        business_types = config.get_business_types()
        business_result_columns = []
        
        for bt_code in target_business_types:
            display_name = business_types[bt_code]['display_name']
            result_col_name = f"{display_name}_替換結果"
            if result_col_name not in header:
                missing_columns.append(result_col_name)
            else:
                business_result_columns.append(result_col_name)
        
        if missing_columns:
            error_msg = f"Excel 缺少必要欄位：{missing_columns}"
            print(f"❌ {error_msg}")
            log_detail(f"錯誤: {error_msg}")
            return None, None, None
        
        log_detail(f"業態替換結果欄位: {business_result_columns}")
        
        return wb, ws, header
        
    except Exception as e:
        error_msg = f"讀取 Excel 檔案失敗：{e}"
        print(f"❌ {error_msg}")
        log_detail(f"錯誤: {error_msg}")
        return None, None, None


def parse_excel_updates(ws, header, config, target_business_types: list, log_detail) -> dict:
    """解析 Excel 中的修正資料，只返回有有效替換的項目"""
    log_detail("開始解析 Excel 修正資料")
    updates = {bt_code: {"po": [], "json": []} for bt_code in target_business_types}
    stats = defaultdict(int)
    
    def get_column_index(name: str) -> int:
        if name not in header:
            raise KeyError(f"Excel 缺少欄位：{name}")
        return header[name]
    
    def get_optional_column_index(name: str) -> int:
        """獲取可選欄位索引，如果不存在返回 -1"""
        return header.get(name, -1)
    
    business_types = config.get_business_types()
    
    # 獲取可選欄位索引
    match_pos_idx = get_optional_column_index("匹配位置")
    category_idx = get_optional_column_index("敏感詞分類")
    
    # 統計變數
    skipped_empty_replacements = 0
    skipped_same_as_original = 0
    valid_updates = 0
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) <= max(header.values()):
            continue
        
        try:
            file_type = row[get_column_index("檔案類型")]
            entry_id = row[get_column_index("項目ID")]
            original_text = row[get_column_index("項目內容")]
            sensitive_word = row[get_column_index("敏感詞")]
            
            if not file_type or not entry_id:
                continue
            
            file_type = str(file_type).lower()
            stats['total_rows'] += 1
            
            # 讀取調試信息（如果存在）
            debug_info = {}
            if match_pos_idx >= 0 and match_pos_idx < len(row) and row[match_pos_idx]:
                debug_info['match_position'] = str(row[match_pos_idx])
            
            if category_idx >= 0 and category_idx < len(row) and row[category_idx]:
                debug_info['category'] = str(row[category_idx])
            
            # 解析多重敏感詞（以逗號分隔）
            if sensitive_word:
                sensitive_words_list = [w.strip() for w in str(sensitive_word).split(',') if w.strip()]
                debug_info['multiple_sensitive_words'] = sensitive_words_list
                log_detail(f"行 {row_num}: 檢測到 {len(sensitive_words_list)} 個敏感詞: {sensitive_words_list}")
            
            # 處理每個目標業態
            has_valid_update = False
            for bt_code in target_business_types:
                display_name = business_types[bt_code]['display_name']
                result_col_name = f"{display_name}_替換結果"
                
                try:
                    new_value = row[get_column_index(result_col_name)]
                except KeyError:
                    continue
                
                # 嚴格的空值檢查，跳過空白值
                if not new_value or not str(new_value).strip():
                    skipped_empty_replacements += 1
                    log_detail(f"行 {row_num}: 跳過空的替換結果 ({display_name})")
                    continue
                
                new_value = str(new_value).strip()
                
                # 安全檢查：跳過與原文相同的替換結果
                if original_text and str(original_text).strip() == new_value:
                    skipped_same_as_original += 1
                    log_detail(f"行 {row_num}: 跳過與原文相同的替換結果 ({display_name})")
                    continue
                
                # 記錄有效的更新
                stats[f'{bt_code}_updates'] += 1
                has_valid_update = True
                
                # 創建更新記錄
                update_record = (str(entry_id), new_value, debug_info)
                
                if file_type == "po":
                    updates[bt_code]["po"].append(update_record)
                elif file_type == "json":
                    updates[bt_code]["json"].append(update_record)
            
            if has_valid_update:
                valid_updates += 1
        
        except Exception as e:
            log_detail(f"錯誤: 第 {row_num} 行處理失敗: {e}")
            continue
    
    # 統計報告
    total_updates = sum(stats[f'{bt_code}_updates'] for bt_code in target_business_types if f'{bt_code}_updates' in stats)
    log_detail(f"解析完成 - 有效更新項目數: {valid_updates}")
    log_detail(f"總更新操作數: {total_updates}")
    log_detail(f"跳過統計 - 空替換結果: {skipped_empty_replacements}, 與原文相同: {skipped_same_as_original}")
    
    # 在控制台顯示關鍵統計
    print(f"   📊 有效項目：{valid_updates} 個，總操作：{total_updates} 次")
    if skipped_empty_replacements > 0 or skipped_same_as_original > 0:
        print(f"   📊 安全跳過：空替換 {skipped_empty_replacements} 個，無變化 {skipped_same_as_original} 個")
    
    return updates


def create_partial_po_file(original_po_path: Path, updates_list: list, output_path: Path, log_detail) -> dict:
    """創建部分 PO 檔案，只包含有修改的項目"""
    result = {"success": False, "updated": 0, "errors": [], "details": []}
    
    if not updates_list:
        log_detail("沒有 PO 更新項目，跳過部分檔案生成")
        return {"success": True, "updated": 0, "errors": [], "details": []}
    
    try:
        # 載入原始 PO 檔案
        original_po = polib.pofile(str(original_po_path))
        
        # 創建新的 PO 檔案，保持原有的元信息
        partial_po = polib.POFile()
        
        # 複製元信息
        partial_po.metadata = original_po.metadata.copy()
        partial_po.header = original_po.header
        
        # 添加部分檔案標識到標題
        if 'POT-Creation-Date' in partial_po.metadata:
            creation_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M%z')
            partial_po.metadata['POT-Creation-Date'] = creation_date
        
        # 添加自定義標題註解
        if partial_po.header:
            partial_po.header += f"\n# 這是部分檔案，只包含有修改的項目\n# 生成時間：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        
        # 處理更新項目
        for update_record in updates_list:
            # 兼容舊格式和新格式
            if len(update_record) == 2:
                msgid, new_msgstr = update_record
                debug_info = {}
            elif len(update_record) == 3:
                msgid, new_msgstr, debug_info = update_record
            else:
                continue
            
            # 在原始檔案中查找對應的項目
            original_entry = original_po.find(msgid)
            if original_entry:
                # 額外的安全檢查
                if original_entry.msgstr == new_msgstr:
                    log_detail(f"PO 跳過: '{msgid}' 內容無變化")
                    continue
                
                # 創建新的條目，保持原有的元信息
                new_entry = polib.POEntry(
                    msgid=original_entry.msgid,
                    msgstr=new_msgstr,
                    msgctxt=original_entry.msgctxt,
                    msgid_plural=original_entry.msgid_plural,
                    msgstr_plural=original_entry.msgstr_plural,
                    obsolete=original_entry.obsolete,
                    encoding=original_entry.encoding,
                    comment=original_entry.comment,
                    tcomment=original_entry.tcomment,
                    occurrences=original_entry.occurrences,
                    flags=original_entry.flags
                )
                
                # 添加處理信息到註解
                if debug_info.get('multiple_sensitive_words'):
                    sensitive_words = ', '.join(debug_info['multiple_sensitive_words'])
                    new_entry.tcomment += f"\n# 敏感詞: {sensitive_words}"
                
                if debug_info.get('category'):
                    new_entry.tcomment += f"\n# 分類: {debug_info['category']}"
                
                partial_po.append(new_entry)
                result["updated"] += 1
                
                # 記錄詳細信息
                detail_msg = f"PO 更新: '{msgid}'"
                if debug_info.get('multiple_sensitive_words'):
                    sensitive_count = len(debug_info['multiple_sensitive_words'])
                    detail_msg += f" [敏感詞:{sensitive_count}個]"
                
                detail_msg += f" → '{new_msgstr[:50]}{'...' if len(new_msgstr) > 50 else ''}'"
                result["details"].append(detail_msg)
                log_detail(detail_msg)
            else:
                error_msg = f"找不到條目：{msgid}"
                result["errors"].append(error_msg)
                log_detail(f"PO 錯誤: {error_msg}")
        
        # 保存部分檔案
        if result["updated"] > 0:
            # 確保輸出目錄存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            partial_po.save(str(output_path))
            log_detail(f"部分 PO 檔案已儲存: {output_path.name}, 包含 {result['updated']} 個條目")
        
        result["success"] = True
        
    except Exception as e:
        error_msg = f"PO 部分檔案生成失敗：{e}"
        result["errors"].append(error_msg)
        log_detail(f"PO 錯誤: {error_msg}")
    
    return result


def create_partial_json_file(original_json_path: Path, updates_list: list, output_path: Path, log_detail) -> dict:
    """創建部分 JSON 檔案，只包含有修改的項目"""
    result = {"success": False, "updated": 0, "errors": [], "details": []}
    
    if not updates_list:
        log_detail("沒有 JSON 更新項目，跳過部分檔案生成")
        return {"success": True, "updated": 0, "errors": [], "details": []}
    
    try:
        # 載入原始 JSON 檔案
        original_data = json.loads(original_json_path.read_text(encoding="utf-8"))
        
        # 創建部分數據結構
        partial_data = {}
        
        # 添加元信息
        partial_data["_metadata"] = {
            "type": "partial_translation",
            "source_file": str(original_json_path),
            "generated_at": datetime.datetime.now().isoformat(),
            "description": "This file contains only modified translation entries"
        }
        
        # 處理更新項目
        for update_record in updates_list:
            # 兼容舊格式和新格式
            if len(update_record) == 2:
                json_path_str, new_value = update_record
                debug_info = {}
            elif len(update_record) == 3:
                json_path_str, new_value, debug_info = update_record
            else:
                continue
            
            # 獲取原始值進行比較
            original_value = get_json_value_by_path(original_data, json_path_str)
            
            # 額外的安全檢查
            if original_value == new_value:
                log_detail(f"JSON 跳過: '{json_path_str}' 內容無變化")
                continue
            
            # 設置新值到部分數據結構中
            if set_json_value_by_path(partial_data, json_path_str, new_value):
                result["updated"] += 1
                
                # 記錄詳細信息
                detail_msg = f"JSON 更新: '{json_path_str}'"
                if debug_info.get('multiple_sensitive_words'):
                    sensitive_count = len(debug_info['multiple_sensitive_words'])
                    detail_msg += f" [敏感詞:{sensitive_count}個]"
                
                detail_msg += f" → '{new_value[:50]}{'...' if len(new_value) > 50 else ''}'"
                result["details"].append(detail_msg)
                log_detail(detail_msg)
            else:
                error_msg = f"無法設置路徑：{json_path_str}"
                result["errors"].append(error_msg)
                log_detail(f"JSON 錯誤: {error_msg}")
        
        # 保存部分檔案
        if result["updated"] > 0:
            # 確保輸出目錄存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            json_content = json.dumps(partial_data, ensure_ascii=False, indent=2)
            output_path.write_text(json_content, encoding="utf-8")
            log_detail(f"部分 JSON 檔案已儲存: {output_path.name}, 包含 {result['updated']} 個條目")
        
        result["success"] = True
        
    except json.JSONDecodeError as e:
        error_msg = f"JSON 格式錯誤：{e}"
        result["errors"].append(error_msg)
        log_detail(f"JSON 錯誤: {error_msg}")
    except Exception as e:
        error_msg = f"JSON 部分檔案生成失敗：{e}"
        result["errors"].append(error_msg)
        log_detail(f"JSON 錯誤: {error_msg}")
    
    return result


def get_json_value_by_path(data: dict, path: str):
    """按路徑獲取 JSON 值"""
    try:
        path_parts = parse_json_path(path)
        current = data
        
        for part_type, part_value in path_parts:
            if part_type == 'key':
                if part_value not in current:
                    return None
                current = current[part_value]
            elif part_type == 'index':
                if not isinstance(current, list) or len(current) <= part_value:
                    return None
                current = current[part_value]
        
        return current
        
    except Exception:
        return None


def parse_json_path(path: str) -> list:
    """解析 JSON 路徑"""
    parts = []
    current = ""
    in_bracket = False
    
    for char in path:
        if char == '[':
            if current:
                parts.append(('key', current))
                current = ""
            in_bracket = True
        elif char == ']':
            if in_bracket and current:
                try:
                    parts.append(('index', int(current)))
                except ValueError:
                    raise ValueError(f"無效的陣列索引：{current}")
                current = ""
            in_bracket = False
        elif char == '.' and not in_bracket:
            if current:
                parts.append(('key', current))
                current = ""
        else:
            current += char
    
    if current:
        parts.append(('key', current))
    
    return parts


def set_json_value_by_path(data: dict, path: str, new_value: str) -> bool:
    """按路徑設置 JSON 值"""
    try:
        path_parts = parse_json_path(path)
        current = data
        
        for i, (part_type, part_value) in enumerate(path_parts):
            is_last = (i == len(path_parts) - 1)
            
            if part_type == 'key':
                if is_last:
                    current[part_value] = new_value
                else:
                    if part_value not in current:
                        next_part_type = path_parts[i + 1][0] if i + 1 < len(path_parts) else 'key'
                        current[part_value] = [] if next_part_type == 'index' else {}
                    current = current[part_value]
            
            elif part_type == 'index':
                if is_last:
                    while len(current) <= part_value:
                        current.append(None)
                    current[part_value] = new_value
                else:
                    while len(current) <= part_value:
                        current.append(None)
                    if current[part_value] is None:
                        next_part_type = path_parts[i + 1][0] if i + 1 < len(path_parts) else 'key'
                        current[part_value] = [] if next_part_type == 'index' else {}
                    current = current[part_value]
        
        return True
        
    except Exception as e:
        return False


def detect_tobemodified_files(config) -> dict:
    """檢測可用的 tobemodified 檔案"""
    available_files = {}
    
    # 檢測輸出目錄中的檔案
    try:
        if hasattr(config, 'get_output_dir'):
            output_dir = config.get_output_dir()
        elif hasattr(config, 'output_dir'):
            output_dir = config.output_dir
        elif hasattr(config, 'get_config'):
            config_data = config.get_config()
            output_dir = Path(config_data.get('output_dir', 'i18n_output'))
        else:
            output_dir = Path('i18n_output')
    except Exception:
        output_dir = Path('i18n_output')
    
    # 使用配置載入器的語言檢測
    try:
        available_languages = config.detect_available_languages()
    except Exception as e:
        print(f"⚠️  語言檢測失敗：{e}")
        available_languages = []
    
    # 檢測標準命名的檔案
    for language in available_languages:
        tobemodified_path = output_dir / f"{language}_tobemodified.xlsx"
        if tobemodified_path.exists():
            available_files[language] = tobemodified_path
    
    # 在當前目錄和輸出目錄中查找額外的 tobemodified 檔案
    for search_dir in [output_dir]:
        if search_dir.exists():
            for file_path in search_dir.glob("*_tobemodified.xlsx"):
                filename = file_path.stem
                if filename.endswith('_tobemodified'):
                    language = filename[:-len('_tobemodified')]
                    
                    # 過濾系統臨時檔案
                    if language.startswith(('~$', '.', '__')):
                        continue
                    
                    if language not in available_files:
                        available_files[language] = file_path

    return available_files


def choose_business_types(config, args) -> list:
    """選擇要處理的業態"""
    if args.business_types:
        if 'all' in args.business_types:
            return list(config.get_business_types().keys())
        return args.business_types
    
    # 互動式選擇
    business_types = config.get_business_types()
    choices = list(business_types.items())
    
    print("\n請選擇要套用修正的業態：")
    for i, (bt_code, bt_config) in enumerate(choices, 1):
        print(f"  {i}) {bt_config['display_name']}")
    print(f"  {len(choices) + 1}) 全部")
    
    while True:
        try:
            opt = input(f"\n輸入選項 (1-{len(choices) + 1})：").strip()
            choice_idx = int(opt) - 1
            
            if choice_idx == len(choices):  # 全部
                return list(business_types.keys())
            elif 0 <= choice_idx < len(choices):
                bt_code = choices[choice_idx][0]
                return [bt_code]
            else:
                print(f"⚠️  請輸入 1-{len(choices) + 1} 之間的數字")
        except (ValueError, KeyboardInterrupt):
            print("\n❌ 使用者取消操作")
            sys.exit(0)


def process_language(config, language: str, target_business_types: list) -> bool:
    """處理單個語言的部分檔案生成"""
    
    # 獲取檔案路徑
    available_files = detect_tobemodified_files(config)
    tobemodified_path = available_files.get(language)
    
    if not tobemodified_path:
        print(f"❌ 找不到 {language} 的 tobemodified 檔案")
        return False
    
    language_files = config.get_language_files(language)
    
    print(f"   來源 Excel：{tobemodified_path.name}")
    
    # 獲取輸出路徑
    try:
        output_paths = config.get_output_paths(language)
        output_dir = output_paths['output_dir']
        timestamp = output_paths['timestamp']
    except Exception:
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        try:
            if hasattr(config, 'get_output_dir'):
                base_output_dir = config.get_output_dir()
            else:
                base_output_dir = Path('i18n_output')
        except Exception:
            base_output_dir = Path('i18n_output')
        
        output_dir = base_output_dir / f"{language}_{timestamp}_partial"
    else:
        # 為部分檔案添加 _partial 後綴
        output_dir = output_dir.parent / f"{output_dir.name}_partial"
    
    # 創建輸出目錄
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 設置日誌
    log_file = output_dir / f"apply_fixes_partial_{timestamp}.log"
    
    def log_detail(message: str):
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(f"{datetime.datetime.now().strftime('%H:%M:%S')} - {message}\n")
    
    log_detail(f"開始處理語言: {language} (部分檔案模式)")
    log_detail(f"目標業態: {', '.join(target_business_types)}")
    log_detail(f"來源檔案: {tobemodified_path}")
    
    # 讀取並驗證 Excel
    wb, ws, header = read_and_validate_xlsx(tobemodified_path, config, target_business_types, log_detail)
    if not wb:
        return False
    
    # 解析修正資料
    updates = parse_excel_updates(ws, header, config, target_business_types, log_detail)
    
    # 處理每個業態
    business_types = config.get_business_types()
    results = {}
    
    for bt_code in target_business_types:
        bt_config = business_types[bt_code]
        display_name = bt_config['display_name']
        suffix = bt_config['suffix']
        
        print(f"   📝 處理 {display_name}...")
        log_detail(f"開始處理業態: {display_name}")
        
        # 生成部分檔案
        result = generate_partial_files(
            config, language, bt_code, updates[bt_code], language_files, output_dir, suffix, log_detail
        )
        
        results[bt_code] = result
        
        if result['success']:
            total_updates = result['po_updated'] + result['json_updated']
            print(f"     ✅ 完成 - PO: {result['po_updated']} 個, JSON: {result['json_updated']} 個")
            log_detail(f"{display_name} 處理完成: 總更新 {total_updates} 個")
        else:
            print(f"     ❌ 失敗")
            log_detail(f"{display_name} 處理失敗")
            
            # 記錄錯誤詳情到日誌
            for error in result.get('errors', []):
                log_detail(f"  錯誤: {error}")
    
    # 生成最終報告
    success_count = sum(1 for r in results.values() if r['success'])
    total_count = len(results)
    total_updates = sum(r['po_updated'] + r['json_updated'] for r in results.values())
    
    print(f"   📊 處理結果：成功 {success_count}/{total_count}，總更新 {total_updates} 個")
    print(f"   📁 輸出目錄：{output_dir}")
    
    log_detail(f"語言 {language} 處理完成: 成功 {success_count}/{total_count} 個業態")
    
    # 生成處理摘要
    generate_summary_report(results, output_dir, timestamp, log_detail)
    
    return success_count > 0


def generate_partial_files(config, language: str, bt_code: str, updates: dict, language_files: dict, output_dir: Path, suffix: str, log_detail) -> dict:
    """生成部分檔案"""
    result = {
        'success': True,
        'po_updated': 0,
        'json_updated': 0,
        'errors': [],
        'details': []
    }
    
    try:
        # 生成部分 PO 檔案
        if 'po_file' in language_files and updates['po']:
            original_po_path = language_files['po_file']
            partial_po_path = output_dir / f"messages{suffix}_partial.po"
            
            po_result = create_partial_po_file(original_po_path, updates['po'], partial_po_path, log_detail)
            result['po_updated'] = po_result['updated']
            result['errors'].extend(po_result['errors'])
            result['details'].extend(po_result.get('details', []))
            if not po_result['success']:
                result['success'] = False
        
        # 生成部分 JSON 檔案
        if 'json_file' in language_files and updates['json']:
            original_json_path = language_files['json_file']
            partial_json_path = output_dir / f"{language}{suffix}_partial.json"
            
            json_result = create_partial_json_file(original_json_path, updates['json'], partial_json_path, log_detail)
            result['json_updated'] = json_result['updated']
            result['errors'].extend(json_result['errors'])
            result['details'].extend(json_result.get('details', []))
            if not json_result['success']:
                result['success'] = False
        
    except Exception as e:
        error_msg = f"生成部分檔案失敗：{e}"
        result['errors'].append(error_msg)
        result['success'] = False
        log_detail(f"錯誤: {error_msg}")
    
    return result


def generate_summary_report(results: dict, output_dir: Path, timestamp: str, log_detail):
    """生成處理摘要報告"""
    summary_file = output_dir / f"partial_processing_summary_{timestamp}.txt"
    
    try:
        with open(summary_file, 'w', encoding='utf-8') as f:
            f.write(f"敏感詞修正部分檔案處理摘要報告\n")
            f.write(f"生成時間：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"{'='*50}\n\n")
            
            total_po_updates = 0
            total_json_updates = 0
            successful_business_types = []
            failed_business_types = []
            
            for bt_code, result in results.items():
                f.write(f"業態：{bt_code}\n")
                f.write(f"狀態：{'成功' if result['success'] else '失敗'}\n")
                f.write(f"PO 更新數量：{result['po_updated']}\n")
                f.write(f"JSON 更新數量：{result['json_updated']}\n")
                
                if result['success']:
                    successful_business_types.append(bt_code)
                    total_po_updates += result['po_updated']
                    total_json_updates += result['json_updated']
                else:
                    failed_business_types.append(bt_code)
                
                if result.get('errors'):
                    f.write(f"錯誤：\n")
                    for error in result['errors']:
                        f.write(f"  - {error}\n")
                
                if result.get('details'):
                    f.write(f"詳細更新記錄：\n")
                    for detail in result['details'][:20]:  # 限制顯示前20條
                        f.write(f"  - {detail}\n")
                            
                    if len(result['details']) > 20:
                        f.write(f"  ... 還有 {len(result['details']) - 20} 條記錄\n")
                
                f.write(f"\n{'-'*30}\n\n")
            
            # 總計統計
            f.write(f"處理總結：\n")
            f.write(f"成功業態：{len(successful_business_types)}\n")
            f.write(f"失敗業態：{len(failed_business_types)}\n")
            f.write(f"總 PO 更新：{total_po_updates}\n")
            f.write(f"總 JSON 更新：{total_json_updates}\n")
            f.write(f"總更新項目：{total_po_updates + total_json_updates}\n")
            
            if successful_business_types:
                f.write(f"\n成功的業態：{', '.join(successful_business_types)}\n")
            
            if failed_business_types:
                f.write(f"失敗的業態：{', '.join(failed_business_types)}\n")
            
            # 部分檔案說明
            f.write(f"\n部分檔案說明：\n")
            f.write(f"- 本次生成的是部分檔案，只包含有修改的翻譯項目\n")
            f.write(f"- PO 檔案：messages_<業態>_partial.po\n")
            f.write(f"- JSON 檔案：<語言>_<業態>_partial.json\n")
            f.write(f"- 部分檔案保持與原檔案相同的結構和層級\n")
            f.write(f"- 可以直接用於 i18n 系統或合併回原檔案\n")
            
            f.write(f"\n使用建議：\n")
            f.write(f"- 檢查生成的部分檔案內容是否正確\n")
            f.write(f"- 確認翻譯項目的層級結構正確\n")
            f.write(f"- 在生產環境使用前進行測試\n")
            f.write(f"- 考慮將部分檔案合併回主檔案\n")
        
        log_detail(f"摘要報告已生成：{summary_file}")
        
    except Exception as e:
        log_detail(f"生成摘要報告失敗：{e}")


def main():
    """主執行函數"""
    print("🚀 開始生成敏感詞修正部分檔案 (v1.0)")
    
    # 載入配置
    config = get_config()
    config.print_config_summary()
    
    # 處理命令列參數
    parser = argparse.ArgumentParser(description='生成敏感詞修正部分檔案')
    parser.add_argument('--language', '-l', 
                       help='指定要處理的語言（若未指定將自動檢測）')
    parser.add_argument('--business-types', '-b',
                       nargs='+',
                       choices=list(config.get_business_types().keys()) + ['all'],
                       help='指定要處理的業態 (可多選，或使用 all)')
    parser.add_argument('--list-files', action='store_true',
                       help='列出所有可用的 tobemodified 檔案')
    
    args = parser.parse_args()
    
    # 檢測可用的 tobemodified 檔案
    available_files = detect_tobemodified_files(config)
    
    if args.list_files:
        print(f"\n📄 可用的 tobemodified 檔案：")
        for lang, filepath in available_files.items():
            print(f"   {lang}: {filepath}")
        return
    
    if not available_files:
        print("❌ 未找到任何 tobemodified 檔案")
        print("請先執行 script_01_generate_xlsx.py 生成檔案")
        sys.exit(1)
    
    # 選擇要處理的語言
    if args.language:
        if args.language not in available_files:
            print(f"❌ 語言 '{args.language}' 的 tobemodified 檔案不存在")
            print(f"可用語言：{list(available_files.keys())}")
            sys.exit(1)
        target_languages = [args.language]
        print(f"\n🌐 將處理指定語言：{args.language}")
    else:
        target_languages = list(available_files.keys())
        print(f"\n🌐 將處理所有語言：{', '.join(target_languages)}")
    
    # 選擇業態
    target_business_types = choose_business_types(config, args)
    
    # 處理每個語言
    success_count = 0
    total_count = len(target_languages)
    
    for language in target_languages:
        print(f"\n{'='*60}")
        print(f"📋 處理語言：{language} (部分檔案模式)")
        
        if process_language(config, language, target_business_types):
            success_count += 1
        else:
            print(f"❌ {language} 處理失敗")
    
    # 最終報告
    print(f"\n🎉 部分檔案生成完畢！")
    print(f"📊 成功處理：{success_count}/{total_count} 個語言")
    print(f"💡 生成的部分檔案特點：")
    print(f"   - 只包含有修改的翻譯項目")
    print(f"   - 保持原有檔案結構和層級")
    print(f"   - 可直接用於 i18n 系統")
    print(f"   - 檔案名包含 '_partial' 標識")
    
    if success_count < total_count:
        sys.exit(1)


if __name__ == "__main__":
    main()