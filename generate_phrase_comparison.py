#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_phrase_comparison.py (v2.3 - 修正路徑結構版本)

生成統一的 phrase_comparison.xlsx 檔案，包含所有語言的敏感詞對照表
支援語言區塊分離，在同一個 Excel 中統一管理所有語言和業態

修正內容：
1. 適配新的路徑結構：JSON 在語言根目錄，PO 在 LC_MESSAGES 子目錄
2. 使用新的 config_loader 方法
"""

import json
import re
import itertools
import sys
import shutil
import datetime
from pathlib import Path
from collections import defaultdict
from config_loader import get_config

try:
    import polib
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
except ImportError as e:
    print(f"❌ 缺少必要套件：{e}")
    print("請執行：pip install polib openpyxl")
    sys.exit(1)


def auto_adjust_column_widths(worksheet, max_width=50):
    """
    自動調整列寬，避免 MergedCell 錯誤
    
    Args:
        worksheet: openpyxl 工作表對象
        max_width: 最大列寬
    """
    try:
        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # 遍歷該列的所有單元格
            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # 跳過 MergedCell
                if isinstance(cell, MergedCell):
                    continue
                    
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # 設置列寬（最小12，最大max_width）
            adjusted_width = min(max(max_length + 4, 12), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        print(f"⚠️  列寬調整發生錯誤（不影響功能）: {e}")


def safe_adjust_column_widths_for_summary(worksheet):
    """
    為總覽工作表安全地調整列寬
    """
    try:
        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                if isinstance(cell, MergedCell):
                    continue
                    
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        print(f"⚠️  總覽工作表列寬調整發生錯誤: {e}")


# 基礎敏感詞字典 - 從 JSON 檔案分析和教育培訓領域經驗整理
BASE_SENSITIVE_WORDS = {
    "學員相關": [
        "學生", "學員", "參與者", "受訓者", "同學", "班級", "組別"
    ],
    "師資相關": [
        "老師", "教師", "講師", "教授", "助教", "指導員", "輔導員"
    ],
    "時間相關": [
        "學期", "學年", "年度", "季度", "月份", "週次", "節次"
    ]
}


def main():
    """主執行函數"""
    print("🚀 開始生成統一的 phrase_comparison.xlsx 檔案")
    
    # 載入配置
    config = get_config()
    config.print_config_summary()
    
    # 檢測可用語言
    available_languages = config.detect_available_languages()
    print(f"\n🌐 將處理 {len(available_languages)} 個語言：{', '.join(available_languages)}")
    
    # 備份配置
    backup_config = config.get_backup_config()
    backup_dir = config.get_backup_dir()
    backup_dir.mkdir(exist_ok=True)
    
    timestamp_format = backup_config.get('timestamp_format', '%Y%m%d_%H%M%S')
    timestamp = datetime.datetime.now().strftime(timestamp_format)
    
    # 獲取統一 Excel 檔案路徑
    excel_path = config.get_comparison_excel_path()
    
    print(f"   目標檔案：{excel_path}")
    
    # 備份現有檔案
    if excel_path.exists():
        backup_filename = f"{excel_path.stem}_{timestamp}{excel_path.suffix}"
        backup_path = backup_dir / backup_filename
        
        shutil.copy2(excel_path, backup_path)
        print(f"   ✅ 已備份現有檔案：{backup_filename}")
    
    # 收集所有語言的敏感詞
    all_language_keywords = {}
    
    for language in available_languages:
        print(f"\n📋 分析語言：{language}")
        try:
            language_files = config.get_language_files(language)
            detected_keywords = detect_sensitive_words(language_files, config, language)
            
            if not detected_keywords:
                print(f"   ⚠️  在 {language} 中未檢測到敏感詞，使用基礎詞彙")
                detected_keywords = BASE_SENSITIVE_WORDS.copy()
            
            all_language_keywords[language] = detected_keywords
            
            total_words = sum(len(words) for words in detected_keywords.values())
            print(f"   📊 檢測到 {total_words} 個敏感詞，{len(detected_keywords)} 個分類")
            for category, words in detected_keywords.items():
                print(f"      {category}: {len(words)} 個詞")
                
        except Exception as e:
            print(f"   ❌ 處理語言 {language} 時發生錯誤：{e}")
            print(f"   ⚠️  將使用基礎詞彙作為後備方案")
            all_language_keywords[language] = BASE_SENSITIVE_WORDS.copy()
    
    # 生成統一 Excel
    generate_unified_excel(config, all_language_keywords, excel_path)
    print(f"\n✅ 統一對照表生成完成：{excel_path}")
    
    # 生成統計報告
    total_languages = len(all_language_keywords)
    total_categories = len(set().union(*[keywords.keys() for keywords in all_language_keywords.values()]))
    total_words = sum(sum(len(words) for words in keywords.values()) for keywords in all_language_keywords.values())
    
    print(f"\n📊 統計報告：")
    print(f"   語言數量：{total_languages}")
    print(f"   分類總數：{total_categories}")
    print(f"   敏感詞總數：{total_words}")
    print(f"   平均每語言：{total_words // total_languages if total_languages else 0} 個敏感詞")


def detect_sensitive_words(language_files: dict, config, language: str) -> dict:
    """
    從語言檔案中檢測敏感詞 - 修正版本，適配新的路徑結構
    
    Args:
        language_files: 語言檔案路徑字典
        config: 配置物件
        language: 語言代碼
    
    Returns:
        檢測到的敏感詞字典 {分類: [詞彙列表]}
    """
    
    # 收集所有文本內容
    all_texts = []
    
    # 讀取 PO 檔案
    if 'po_file' in language_files:
        try:
            po_file_path = language_files['po_file']
            print(f"      📖 讀取 PO 檔案：{po_file_path}")
            po_file = polib.pofile(str(po_file_path))
            for entry in po_file:
                if entry.msgid:
                    all_texts.append(entry.msgid)
                if entry.msgstr:
                    all_texts.append(entry.msgstr)
            print(f"      ✅ PO 檔案：{len(po_file)} 個條目")
        except Exception as e:
            print(f"      ⚠️  讀取 PO 檔案失敗：{e}")
    
    # 讀取 JSON 檔案
    if 'json_file' in language_files:
        try:
            json_file_path = language_files['json_file']
            print(f"      📖 讀取 JSON 檔案：{json_file_path}")
            with open(json_file_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            def extract_json_values(obj):
                """遞迴提取 JSON 中的所有字符串值"""
                if isinstance(obj, dict):
                    for value in obj.values():
                        yield from extract_json_values(value)
                elif isinstance(obj, list):
                    for item in obj:
                        yield from extract_json_values(item)
                elif isinstance(obj, str):
                    yield obj
            
            json_texts = list(extract_json_values(json_data))
            all_texts.extend(json_texts)
            print(f"      ✅ JSON 檔案：{len(json_texts)} 個文本")
            
        except Exception as e:
            print(f"      ⚠️  讀取 JSON 檔案失敗：{e}")
    
    if not all_texts:
        print(f"      ⚠️  無法從檔案中提取文本內容")
        return {}
    
    # 檢測敏感詞
    print(f"      🔍 從 {len(all_texts)} 個文本條目中檢測敏感詞...")
    
    detected_words = defaultdict(set)
    detection_config = config.get_keyword_detection_config()
    case_sensitive = detection_config.get('case_sensitive', False)
    
    # 合併所有文本
    combined_text = ' '.join(all_texts)
    if not case_sensitive:
        combined_text = combined_text.lower()
    
    # 對每個基礎分類的敏感詞進行檢測
    for category, base_words in BASE_SENSITIVE_WORDS.items():
        for word in base_words:
            search_word = word.lower() if not case_sensitive else word
            if search_word in combined_text:
                detected_words[category].add(word)
    
    # 轉換為普通字典，並按原始順序排列
    result = {}
    for category, words in detected_words.items():
        # 保持與基礎詞典相同的順序
        ordered_words = []
        for base_word in BASE_SENSITIVE_WORDS[category]:
            if base_word in words:
                ordered_words.append(base_word)
        if ordered_words:
            result[category] = ordered_words
    
    return result


def generate_unified_excel(config, all_language_keywords: dict, output_path: Path):
    """
    生成統一的 phrase_comparison Excel 檔案 - 語言獨立橫向分區塊版本
    
    Args:
        config: 配置物件
        all_language_keywords: 所有語言的敏感詞字典
        output_path: 輸出檔案路徑
    """
    
    # 創建工作簿
    wb = Workbook()
    
    # 設置主工作表
    ws = wb.active
    excel_config = config.get_excel_config()
    ws.title = excel_config.get('worksheets', {}).get('comparison', 'phrase_comparison')
    
    # 樣式設定
    styling = excel_config.get('styling', {})
    language_header_color = styling.get('language_header_color', '4472C4')
    category_header_color = styling.get('category_header_color', '70AD47')
    business_header_color = styling.get('business_header_color', 'FFC000')
    data_row_color = styling.get('data_row_color', 'F2F2F2')
    
    # 字體和邊框樣式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    language_font = Font(bold=True, color="FFFFFF", size=14)
    business_font = Font(bold=True, color="FFFFFF", size=10)
    data_font = Font(size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    
    # 建立語言獨立的橫向結構
    business_types = config.get_business_types()
    
    # 橫向配置
    horizontal_config = excel_config.get('horizontal_layout', {})
    block_separator = horizontal_config.get('block_separator_columns', 1)
    
    # 計算每個語言區塊的寬度：敏感詞類型 + 敏感詞 + 業態數量
    block_width = 2 + len(business_types)  # 2 是基礎列數
    
    # 寫入總標題（第1行）
    total_columns = len(all_language_keywords) * block_width + (len(all_language_keywords) - 1) * block_separator
    
    ws.merge_cells(f'A1:{get_column_letter(total_columns)}1')
    title_cell = ws['A1']
    title_cell.value = "多語言敏感詞對照表（語言獨立橫向分區塊版）"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 為每個語言創建獨立區塊
    current_col = 1
    
    for lang_index, (language, keywords_dict) in enumerate(all_language_keywords.items()):
        block_start_col = current_col
        block_end_col = current_col + block_width - 1
        
        # 語言標題（第2行，跨越整個區塊）
        ws.merge_cells(f'{get_column_letter(block_start_col)}2:{get_column_letter(block_end_col)}2')
        lang_cell = ws.cell(row=2, column=block_start_col, value=f"{language}")
        lang_cell.font = language_font
        lang_cell.fill = PatternFill(start_color=language_header_color, end_color=language_header_color, fill_type="solid")
        lang_cell.alignment = Alignment(horizontal="center", vertical="center")
        lang_cell.border = thick_border
        
        # 區塊內標題列（第3行）
        block_headers = ["敏感詞類型", "敏感詞"]
        for bt_code, bt_config in business_types.items():
            block_headers.append(bt_config['display_name'])
        
        for i, header in enumerate(block_headers):
            col = block_start_col + i
            cell = ws.cell(row=3, column=col, value=header)
            
            if i < 2:  # 基礎列
                cell.font = header_font
                cell.fill = PatternFill(start_color=category_header_color, end_color=category_header_color, fill_type="solid")
            else:  # 業態列
                cell.font = business_font
                cell.fill = PatternFill(start_color=business_header_color, end_color=business_header_color, fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # 寫入該語言的敏感詞資料（從第4行開始）
        current_row = 4
        
        for category, keywords in keywords_dict.items():
            for keyword_index, keyword in enumerate(keywords):
                # 敏感詞類型和敏感詞
                ws.cell(row=current_row, column=block_start_col, value=category if keyword_index == 0 else "")
                ws.cell(row=current_row, column=block_start_col + 1, value=keyword)
                
                # 為每個業態添加空白方案欄位
                for bt_index in range(len(business_types)):
                    col = block_start_col + 2 + bt_index
                    cell = ws.cell(row=current_row, column=col, value="")
                    cell.border = thin_border
                    # 設置背景色（奇偶行）
                    if current_row % 2 == 0:
                        cell.fill = PatternFill(start_color=data_row_color, end_color=data_row_color, fill_type="solid")
                
                # 設置基礎列的樣式
                for base_col_offset in [0, 1]:
                    col = block_start_col + base_col_offset
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if current_row % 2 == 0:
                        cell.fill = PatternFill(start_color=data_row_color, end_color=data_row_color, fill_type="solid")
                
                current_row += 1
        
        # 移動到下個語言區塊
        current_col = block_end_col + 1 + block_separator
    
    # 自動調整欄寬
    auto_adjust_column_widths(ws, max_width=25)
    
    # 創建總覽工作表
    create_summary_worksheet(wb, config, all_language_keywords)
    
    # 確保輸出目錄存在
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 保存檔案
    wb.save(output_path)
    
    total_languages = len(all_language_keywords)
    total_keywords = sum(sum(len(words) for words in keywords.values()) for keywords in all_language_keywords.values())
    print(f"      📊 Excel 統計：{total_languages} 個語言，每個語言獨立區塊")
    print(f"      📐 總敏感詞數：{total_keywords} 個")
    print(f"      📏 表格寬度：{total_columns} 列")


def create_summary_worksheet(wb, config, all_language_keywords: dict):
    """
    創建語言總覽工作表
    
    Args:
        wb: 工作簿物件
        config: 配置物件
        all_language_keywords: 所有語言的敏感詞字典
    """
    
    # 創建總覽工作表
    excel_config = config.get_excel_config()
    summary_sheet_name = excel_config.get('worksheets', {}).get('summary', '語言總覽')
    summary_ws = wb.create_sheet(title=summary_sheet_name)
    
    # 樣式設定
    header_font = Font(bold=True, color="FFFFFF", size=12)
    data_font = Font(size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 標題
    summary_ws['A1'] = "語言總覽統計"
    title_cell = summary_ws['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # 統計表頭
    headers = ["語言", "檔案類型", "分類數量", "敏感詞數量", "備註"]
    for col_num, header in enumerate(headers, 1):
        cell = summary_ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # 統計資料
    current_row = 4
    business_types = config.get_business_types()
    
    for language, keywords_dict in all_language_keywords.items():
        # 獲取語言檔案資訊
        try:
            language_files = config.get_language_files(language)
            file_types = []
            if 'po_file' in language_files:
                file_types.append('PO')
            if 'json_file' in language_files:
                file_types.append('JSON')
        except Exception:
            file_types = []
        
        file_type_str = '+'.join(file_types) if file_types else "無檔案"
        category_count = len(keywords_dict)
        keyword_count = sum(len(words) for words in keywords_dict.values())
        
        # 備註資訊
        notes = []
        if keyword_count == 0:
            notes.append("無敏感詞")
        elif keyword_count < 20:
            notes.append("敏感詞較少")
        
        row_data = [
            language,
            file_type_str,
            category_count,
            keyword_count,
            '；'.join(notes) if notes else "正常"
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=current_row, column=col_num, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_num != 5 else "left", vertical="center")
        
        current_row += 1
    
    # 總計行
    total_languages = len(all_language_keywords)
    total_categories = len(set().union(*[keywords.keys() for keywords in all_language_keywords.values()]))
    total_keywords = sum(sum(len(words) for words in keywords.values()) for keywords in all_language_keywords.values())
    
    total_row_data = [
        f"總計 ({total_languages} 個語言)",
        "-",
        total_categories,
        total_keywords,
        f"平均每語言 {total_keywords // total_languages if total_languages else 0} 個敏感詞"
    ]
    
    for col_num, value in enumerate(total_row_data, 1):
        cell = summary_ws.cell(row=current_row, column=col_num, value=value)
        cell.font = Font(bold=True, size=10)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_num != 5 else "left", vertical="center")
    
    # 業態資訊
    current_row += 3
    summary_ws.cell(row=current_row, column=1, value="支援的業態：").font = Font(bold=True)
    current_row += 1
    
    for bt_code, bt_config in business_types.items():
        summary_ws.cell(row=current_row, column=1, value=f"• {bt_config['display_name']}")
        summary_ws.cell(row=current_row, column=2, value=bt_config['description'])
        current_row += 1
    
    # 使用說明
    current_row += 2
    summary_ws.cell(row=current_row, column=1, value="使用說明：").font = Font(bold=True)
    current_row += 1
    
    instructions = [
        "1. 在「phrase_comparison」工作表中編輯各業態的對應方案",
        "2. 空白欄位表示使用原始敏感詞，無需替換",
        "3. 編輯完成後，執行 script_01_generate_xlsx.py 生成待修正清單",
        "4. 最後執行 script_02_apply_fixes.py 套用修正結果"
    ]
    
    for instruction in instructions:
        summary_ws.cell(row=current_row, column=1, value=instruction)
        current_row += 1
    
    # 自動調整欄寬（使用安全方法）
    safe_adjust_column_widths_for_summary(summary_ws)


def test_detection():
    """測試敏感詞檢測功能"""
    print("🧪 測試敏感詞檢測功能...")
    
    # 創建測試文本
    test_texts = [
        "學生成績管理系統中的課程資料",
        "教師可以查看班級學員的學習進度",
        "培訓機構需要統計學員的出席率",
        "系統管理員負責維護用戶帳號權限"
    ]
    
    # 模擬檢測
    combined_text = ' '.join(test_texts).lower()
    
    detected = defaultdict(list)
    for category, words in BASE_SENSITIVE_WORDS.items():
        for word in words:
            if word in combined_text:
                detected[category].append(word)
    
    print("檢測結果：")
    for category, words in detected.items():
        if words:
            print(f"  {category}: {', '.join(words)}")
    
    print(f"\n總計檢測到 {sum(len(words) for words in detected.values())} 個敏感詞")
    return dict(detected)


def extract_keywords_from_json(json_file_path: str) -> dict:
    """
    從 JSON 檔案中提取敏感詞（備用功能）
    
    Args:
        json_file_path: JSON 檔案路徑
    
    Returns:
        提取的敏感詞字典
    """
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 這裡可以根據 JSON 的具體結構來提取敏感詞
        # 目前使用預設的基礎詞典
        return BASE_SENSITIVE_WORDS.copy()
        
    except Exception as e:
        print(f"⚠️  從 JSON 檔案提取敏感詞失敗：{e}")
        return BASE_SENSITIVE_WORDS.copy()


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='生成統一的 phrase_comparison Excel 檔案')
    parser.add_argument('--test', action='store_true', help='執行檢測測試')
    parser.add_argument('--extract-json', help='從指定 JSON 檔案提取敏感詞')
    
    args = parser.parse_args()
    
    if args.test:
        test_detection()
    elif args.extract_json:
        # 從 JSON 檔案提取敏感詞的功能
        if Path(args.extract_json).exists():
            extracted = extract_keywords_from_json(args.extract_json)
            print(f"從 {args.extract_json} 提取的敏感詞：")
            for category, words in extracted.items():
                print(f"  {category}: {len(words)} 個詞")
        else:
            print(f"❌ JSON 檔案不存在：{args.extract_json}")
    else:
        # 正常執行
        main()